"""
03. 도메인사전 마이그레이션
엑셀 → meta.std_domain (614건)
선행: 01_migrate_domain_group.py (도메인그룹 필수)
"""
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from config import DB_CONFIG, DOMAIN_FILE, BATCH_SIZE


def clean_str(val):
    if val is None or str(val).strip() == "":
        return None
    return str(val).strip()


def parse_int(val):
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def parse_bool(val):
    if val is None or val == "":
        return False
    return str(val).strip().upper() == "Y"


def migrate():
    print("=" * 60)
    print("03. 도메인사전 마이그레이션 시작")
    print("=" * 60)

    wb = load_workbook(DOMAIN_FILE, read_only=False)
    ws = wb.active

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # 도메인그룹 매핑 캐시
    cur.execute("SELECT group_id, group_name FROM meta.std_domain_group")
    group_map = {name: gid for gid, name in cur.fetchall()}
    print(f"  도메인그룹 {len(group_map)}건 로드")

    success = 0
    errors = 0
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            group_name = str(row[0]).strip()
            group_id = group_map.get(group_name)
            if not group_id:
                errors += 1
                if errors <= 5:
                    print(f"  [경고] 행 {i}: 도메인그룹 '{group_name}' 없음, 건너뜀")
                continue

            batch.append((
                group_id,
                str(row[1]).strip(),           # domain_name
                str(row[2]).strip(),           # domain_logical_name
                str(row[3]).strip() if row[3] else "VARCHAR",  # data_type
                parse_int(row[4]),             # data_length
                parse_int(row[5]),             # data_scale
                parse_bool(row[6]),            # is_personal_info
                parse_bool(row[7]),            # is_encrypted
                clean_str(row[8]),             # scramble_type
                clean_str(row[9]),             # description
            ))
            success += 1

            if len(batch) >= BATCH_SIZE:
                execute_values(cur, """
                    INSERT INTO meta.std_domain
                    (group_id, domain_name, domain_logical_name, data_type, data_length, data_scale,
                     is_personal_info, is_encrypted, scramble_type, description)
                    VALUES %s
                    ON CONFLICT DO NOTHING
                """, batch)
                conn.commit()
                batch = []
        except Exception as e:
            errors += 1
            print(f"  [오류] 행 {i}: {e}")

    if batch:
        execute_values(cur, """
            INSERT INTO meta.std_domain
            (group_id, domain_name, domain_logical_name, data_type, data_length, data_scale,
             is_personal_info, is_encrypted, scramble_type, description)
            VALUES %s
            ON CONFLICT DO NOTHING
        """, batch)
        conn.commit()

    cur.close()
    conn.close()
    wb.close()

    print(f"  완료: {success}건 성공, {errors}건 실패")
    return success


if __name__ == "__main__":
    migrate()
