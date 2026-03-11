"""
02. 단어사전 마이그레이션
엑셀 → meta.std_word (6,291건)
"""
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from config import DB_CONFIG, WORD_FILE, BATCH_SIZE


def clean_str(val):
    """빈문자열 → None"""
    if val is None or str(val).strip() == "":
        return None
    return str(val).strip()


def parse_bool(val):
    """Y/N → boolean"""
    if val is None or val == "":
        return False
    return str(val).strip().upper() == "Y"


def migrate():
    print("=" * 60)
    print("02. 단어사전 마이그레이션 시작")
    print("=" * 60)

    wb = load_workbook(WORD_FILE, read_only=False)
    ws = wb.active

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    success = 0
    errors = 0
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            batch.append((
                str(row[0]).strip(),        # logical_name
                str(row[1]).strip(),        # physical_name
                clean_str(row[2]),          # physical_desc
                parse_bool(row[3]),         # is_class_word
                clean_str(row[4]),          # synonym
                clean_str(row[5]),          # description
            ))
            success += 1

            if len(batch) >= BATCH_SIZE:
                execute_values(cur, """
                    INSERT INTO meta.std_word (logical_name, physical_name, physical_desc, is_class_word, synonym, description)
                    VALUES %s
                    ON CONFLICT DO NOTHING
                """, batch)
                conn.commit()
                print(f"  진행: {success}건...")
                batch = []
        except Exception as e:
            errors += 1
            print(f"  [오류] 행 {i}: {e}")

    if batch:
        execute_values(cur, """
            INSERT INTO meta.std_word (logical_name, physical_name, physical_desc, is_class_word, synonym, description)
            VALUES %s
            ON CONFLICT DO NOTHING
        """, batch)
        conn.commit()

    cur.close()
    conn.close()
    wb.close()

    print(f"  완료: {success}건 성공, {errors}건 실패")
    return success


if __name__ == "__main__":
    migrate()
