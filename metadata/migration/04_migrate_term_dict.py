"""
04. 용어사전 마이그레이션
엑셀 → meta.std_term (41,359건)
선행: 01_migrate_domain_group.py
"""
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from config import DB_CONFIG, TERM_FILE, BATCH_SIZE


def clean_str(val):
    if val is None or str(val).strip() == "":
        return None
    return str(val).strip()


def parse_int(val):
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def parse_bool(val):
    if val is None or val == "":
        return False
    return str(val).strip().upper() == "Y"


def migrate():
    print("=" * 60)
    print("04. 용어사전 마이그레이션 시작")
    print("=" * 60)

    wb = load_workbook(TERM_FILE, read_only=False)
    ws = wb.active

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # 도메인그룹 매핑
    cur.execute("SELECT group_id, group_name FROM meta.std_domain_group")
    group_map = {name: gid for gid, name in cur.fetchall()}
    print(f"  도메인그룹 {len(group_map)}건 로드")

    success = 0
    errors = 0
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            domain_group_name = clean_str(row[5])
            domain_group_id = group_map.get(domain_group_name) if domain_group_name else None

            batch.append((
                str(row[0]).strip(),           # logical_name
                str(row[1]).strip(),           # physical_name
                clean_str(row[2]),             # english_name
                clean_str(row[3]),             # domain_logical_name
                clean_str(row[4]),             # domain_abbr
                domain_group_id,               # domain_group_id
                clean_str(row[6]),             # data_type
                parse_int(row[7]),             # data_length
                parse_int(row[8]),             # data_scale
                parse_bool(row[9]),            # is_personal_info
                parse_bool(row[10]),           # is_encrypted
                clean_str(row[11]),            # scramble_type
                clean_str(row[12]),            # description
            ))
            success += 1

            if len(batch) >= BATCH_SIZE:
                execute_values(cur, """
                    INSERT INTO meta.std_term
                    (logical_name, physical_name, english_name, domain_logical_name, domain_abbr,
                     domain_group_id, data_type, data_length, data_scale,
                     is_personal_info, is_encrypted, scramble_type, description)
                    VALUES %s
                    ON CONFLICT DO NOTHING
                """, batch)
                conn.commit()
                print(f"  진행: {success}건...")
                batch = []
        except Exception as e:
            errors += 1
            if errors <= 10:
                print(f"  [오류] 행 {i}: {e}")

    if batch:
        execute_values(cur, """
            INSERT INTO meta.std_term
            (logical_name, physical_name, english_name, domain_logical_name, domain_abbr,
             domain_group_id, data_type, data_length, data_scale,
             is_personal_info, is_encrypted, scramble_type, description)
            VALUES %s
            ON CONFLICT DO NOTHING
        """, batch)
        conn.commit()

    cur.close()
    conn.close()
    wb.close()

    print(f"  완료: {success}건 성공, {errors}건 실패")
    return success


if __name__ == "__main__":
    migrate()
