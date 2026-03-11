"""
01. 도메인그룹 마이그레이션
도메인사전 엑셀에서 도메인그룹명을 추출하여 meta.std_domain_group에 적재
(이미 스키마 SQL에 INSERT 되어 있지만, 엑셀에서 추가 그룹 확인용)
"""
import psycopg2
from openpyxl import load_workbook
from config import DB_CONFIG, DOMAIN_FILE


def migrate():
    print("=" * 60)
    print("01. 도메인그룹 마이그레이션 시작")
    print("=" * 60)

    wb = load_workbook(DOMAIN_FILE, read_only=False)
    ws = wb.active

    # 도메인그룹명 추출 (중복 제거)
    group_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            group_names.add(str(row[0]).strip())

    wb.close()
    print(f"  엑셀에서 {len(group_names)}개 도메인그룹 추출: {sorted(group_names)}")

    # DB 적재
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    inserted = 0
    for i, name in enumerate(sorted(group_names), 1):
        try:
            cur.execute("""
                INSERT INTO meta.std_domain_group (group_name, sort_order)
                VALUES (%s, %s)
                ON CONFLICT (group_name) DO NOTHING
            """, (name, i))
            if cur.rowcount > 0:
                inserted += 1
        except Exception as e:
            print(f"  [오류] '{name}': {e}")

    conn.commit()
    cur.close()
    conn.close()

    print(f"  완료: {inserted}건 신규 적재 (기존 데이터와 중복 제외)")
    return inserted


if __name__ == "__main__":
    migrate()
