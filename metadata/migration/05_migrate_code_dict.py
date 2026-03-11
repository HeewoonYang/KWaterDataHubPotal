"""
05. 코드사전 마이그레이션
엑셀 → meta.std_code_group + meta.std_code (139,258건, 2시트)
"""
import re
import psycopg2
from openpyxl import load_workbook
from config import DB_CONFIG, CODE_FILE, BATCH_SIZE


def clean_str(val):
    if val is None or str(val).strip() == "":
        return None
    return str(val).strip()


def parse_int(val):
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def parse_code_group(val):
    """'ADT(감사)' → ('ADT', '감사')"""
    if not val:
        return None, None
    s = str(val).strip()
    match = re.match(r'^([A-Za-z0-9]+)\((.+)\)$', s)
    if match:
        return match.group(1), match.group(2)
    return s, None


def parse_date(val):
    """날짜 문자열 또는 datetime → 'YYYY-MM-DD' 문자열"""
    if val is None or val == "":
        return None
    try:
        from datetime import datetime
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        if len(s) >= 10:
            return s[:10]
        return None
    except:
        return None


def migrate():
    print("=" * 60)
    print("05. 코드사전 마이그레이션 시작")
    print("=" * 60)

    wb = load_workbook(CODE_FILE, read_only=False)
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    total_groups = 0
    total_codes = 0
    errors = 0

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        print(f"\n  --- Sheet {sheet_idx + 1}: {sheet_name} ---")

        # 코드그룹 캐시 (code_id → group_id)
        group_cache = {}
        code_batch = []
        sheet_codes = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            try:
                # 코드그룹 정보 (처음 나올 때 INSERT)
                sys_prefix, sys_name = parse_code_group(row[0])
                code_id_val = clean_str(row[6])

                if not code_id_val:
                    continue

                if code_id_val not in group_cache:
                    # 코드그룹 UPSERT
                    cur.execute("""
                        INSERT INTO meta.std_code_group
                        (system_prefix, system_name, logical_name, physical_name, code_desc, code_type, data_length, code_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (code_id) DO UPDATE SET system_name = EXCLUDED.system_name
                        RETURNING group_id
                    """, (
                        sys_prefix,
                        sys_name,
                        str(row[1]).strip() if row[1] else "",   # logical_name
                        str(row[2]).strip() if row[2] else "",   # physical_name
                        clean_str(row[3]),                        # code_desc
                        clean_str(row[4]),                        # code_type
                        parse_int(row[5]),                        # data_length
                        code_id_val,                              # code_id
                    ))
                    result = cur.fetchone()
                    group_cache[code_id_val] = result[0]
                    total_groups += 1
                    conn.commit()

                group_id = group_cache[code_id_val]

                # 코드값이 있는 경우만 코드 적재
                code_value = clean_str(row[7])
                code_name = clean_str(row[8])
                if code_value is not None and code_name is not None:
                    code_batch.append((
                        group_id,
                        str(code_value),
                        str(code_name),
                        parse_int(row[9]),           # sort_order
                        clean_str(row[10]),           # parent_code_name
                        clean_str(row[11]),           # parent_code_value
                        clean_str(row[12]),           # description
                        parse_date(row[13]),          # effective_from
                        parse_date(row[14]),          # effective_to
                    ))
                    sheet_codes += 1

                if len(code_batch) >= BATCH_SIZE:
                    cur.executemany("""
                        INSERT INTO meta.std_code
                        (group_id, code_value, code_name, sort_order,
                         parent_code_name, parent_code_value, description,
                         effective_from, effective_to)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (group_id, code_value) DO NOTHING
                    """, code_batch)
                    conn.commit()
                    print(f"    진행: {sheet_codes}건...")
                    code_batch = []

            except Exception as e:
                errors += 1
                if errors <= 20:
                    print(f"    [오류] 행 {i}: {e}")

        # 잔여 배치 처리
        if code_batch:
            cur.executemany("""
                INSERT INTO meta.std_code
                (group_id, code_value, code_name, sort_order,
                 parent_code_name, parent_code_value, description,
                 effective_from, effective_to)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (group_id, code_value) DO NOTHING
            """, code_batch)
            conn.commit()

        total_codes += sheet_codes
        print(f"    Sheet 완료: 코드 {sheet_codes}건")

    cur.close()
    conn.close()
    wb.close()

    print(f"\n  === 전체 완료 ===")
    print(f"  코드그룹: {total_groups}건")
    print(f"  코드값: {total_codes}건")
    print(f"  오류: {errors}건")
    return total_codes


if __name__ == "__main__":
    migrate()
