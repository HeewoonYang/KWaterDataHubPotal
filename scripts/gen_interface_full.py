# -*- coding: utf-8 -*-
"""인터페이스설계서 전체 생성 스크립트 (20건)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = r'D:\00_수공프로젝트\20260224_설계단계_산출물\산출물 템플릿_데이터허브(손이사님 제공)\4. 데이터허브\2. 설계\(템플릿)KWDP-DH-DS-17-인터페이스설계서-v0.x.xlsx'
DST = r'D:\00_수공프로젝트\20260224_설계단계_산출물\산출물 템플릿_데이터허브(손이사님 제공)\4. 데이터허브\2. 설계\KWDP-DH-DS-17-인터페이스설계서-v0.2.xlsx'

wb = openpyxl.load_workbook(SRC)
ws = wb.worksheets[2]

thin = Side(style='thin')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
font_data = Font(name='맑은 고딕', size=10)
font_bold = Font(name='맑은 고딕', size=10, bold=True)
fill_gray = PatternFill('solid', fgColor='F2F2F2')
align_cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_lc = Alignment(horizontal='left', vertical='center', wrap_text=True)


def set_cell(ws, row, col, value, bold=False, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_bold if bold else font_data
    c.border = border_all
    c.alignment = align if align else align_cc
    if fill:
        c.fill = fill


def set_empty_border(ws, row, col, align=None):
    c = ws.cell(row=row, column=col)
    c.font = font_data
    c.border = border_all
    c.alignment = align if align else align_cc


def write_interface(ws, start_row, data):
    total_rows = sum(len(s['fields']) for s in data['sections'])
    end_row = start_row + total_rows - 1

    basic_cols = [
        (1, data['area']), (2, data['if_id']), (3, data['if_name']),
        (4, data['req_sys']), (5, data['res_sys']), (6, data['in_out']),
        (7, data['cycle']), (8, data['method'])
    ]
    for col, val in basic_cols:
        al = align_lc if col == 3 else align_cc
        set_cell(ws, start_row, col, val, align=al)
        if total_rows > 1:
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row, end_column=col)
            for r in range(start_row + 1, end_row + 1):
                set_empty_border(ws, r, col, al)

    cur_row = start_row
    k_groups = []
    cur_k_dir = None
    cur_k_start = None

    for sec in data['sections']:
        sec_start = cur_row
        sec_end = cur_row + len(sec['fields']) - 1

        set_cell(ws, sec_start, 9, sec['i_type'], bold=True, fill=fill_gray)
        if len(sec['fields']) > 1:
            ws.merge_cells(start_row=sec_start, start_column=9,
                           end_row=sec_end, end_column=9)
            for r in range(sec_start + 1, sec_end + 1):
                set_empty_border(ws, r, 9)

        set_cell(ws, sec_start, 10, sec['j_name'], align=align_lc)
        if len(sec['fields']) > 1:
            ws.merge_cells(start_row=sec_start, start_column=10,
                           end_row=sec_end, end_column=10)
            for r in range(sec_start + 1, sec_end + 1):
                set_empty_border(ws, r, 10, align_lc)

        actual_k = sec['k_dir']
        for i, field in enumerate(sec['fields']):
            r = cur_row + i
            if cur_k_dir is None:
                cur_k_dir = actual_k
                cur_k_start = r
            elif actual_k != cur_k_dir:
                k_groups.append((cur_k_start, r - 1, cur_k_dir))
                cur_k_dir = actual_k
                cur_k_start = r

            set_cell(ws, r, 12, field[0], align=align_lc)
            set_cell(ws, r, 13, field[1], align=align_lc)
            set_cell(ws, r, 14, field[2], align=align_lc)
            set_cell(ws, r, 15, field[3])
            set_cell(ws, r, 16, field[4])
            set_cell(ws, r, 17, field[5])
            set_cell(ws, r, 18, field[6] if len(field) > 6 else '', align=align_lc)
            set_cell(ws, r, 19, field[7] if len(field) > 7 else '', align=align_lc)

        cur_row = sec_end + 1

    if cur_k_start is not None:
        k_groups.append((cur_k_start, end_row, cur_k_dir))

    for ks, ke, kd in k_groups:
        set_cell(ws, ks, 11, kd, bold=True, fill=fill_gray)
        if ke > ks:
            ws.merge_cells(start_row=ks, start_column=11,
                           end_row=ke, end_column=11)
            for r in range(ks + 1, ke + 1):
                set_empty_border(ws, r, 11)

    return end_row + 1


# ============================================================
# Batch 1: 시스템관리 (3건)
# ============================================================
batch_sy = [
    {
        'area': '시스템관리', 'if_id': 'IF-SY-001',
        'if_name': 'ERP 인사정보 동기화',
        'req_sys': '데이터허브포털', 'res_sys': 'K-water ERP(SAP)',
        'in_out': '내부', 'cycle': '일 1회', 'method': 'DB',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'ERP 인사정보 동기화', 'k_dir': '송신',
             'fields': [
                 ('동기화유형', 'sync_type', 'VARCHAR', 10, 'N', 'N', '', '전체/증분/CDC'),
                 ('대상구분', 'target_type', 'VARCHAR', 10, 'N', 'N', '', '인사/조직/겸직'),
                 ('요청일시', 'req_date', 'DATETIME', 23, 'N', 'N', '', ''),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_SYS_USER', 'k_dir': '수신',
             'fields': [
                 ('사번', 'emp_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK, PREFIX("KW")'),
                 ('이름', 'user_name', 'VARCHAR', 50, 'N', 'N', '', ''),
                 ('부서코드', 'dept_code', 'VARCHAR', 20, 'N', 'N', '', 'LOOKUP(ORG_MAP)'),
                 ('직급', 'position', 'VARCHAR', 20, 'Y', 'N', '', 'LOOKUP(JOB_MAP)'),
                 ('이메일', 'email', 'VARCHAR', 100, 'Y', 'N', '', 'LOWER()'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '-', 'k_dir': '수신',
             'fields': [
                 ('상태', 'status', 'CHAR', 1, 'N', 'N', 'A', 'MAP(A→활성,T→퇴직)'),
                 ('처리건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
                 ('결과코드', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '시스템관리', 'if_id': 'IF-SY-002',
        'if_name': 'SSO 통합인증 사용자 연계',
        'req_sys': '데이터허브포털', 'res_sys': 'K-water SSO 인증서버',
        'in_out': '내부', 'cycle': '실시간', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '로그인 화면', 'k_dir': '송신',
             'fields': [
                 ('SAML요청', 'saml_request', 'TEXT', '', 'N', 'N', '', 'Base64 인코딩'),
                 ('리턴URL', 'relay_state', 'VARCHAR', 500, 'Y', 'N', '', '인증 후 리턴 주소'),
                 ('인증방식', 'auth_method', 'CHAR', 3, 'N', 'N', 'SSO', 'SSO/PWD'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_SYS_SESSION', 'k_dir': '수신',
             'fields': [
                 ('사번', 'emp_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                 ('사용자명', 'user_name', 'VARCHAR', 50, 'N', 'N', '', ''),
                 ('부서코드', 'dept_code', 'VARCHAR', 20, 'N', 'N', '', ''),
                 ('역할코드', 'role_code', 'VARCHAR', 20, 'N', 'N', '', '관리자/일반/엔지니어'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'https://sso.kwater.or.kr/auth/saml', 'k_dir': '수신',
             'fields': [
                 ('인증토큰', 'auth_token', 'VARCHAR', 500, 'N', 'N', '', 'JWT 토큰'),
                 ('만료일시', 'token_expire', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('결과코드', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '시스템관리', 'if_id': 'IF-SY-003',
        'if_name': 'API 게이트웨이 인증/인가',
        'req_sys': '외부기관/시스템', 'res_sys': '데이터허브포털',
        'in_out': '외부', 'cycle': '실시간', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '연계인터페이스 모니터링', 'k_dir': '송신',
             'fields': [
                 ('API키', 'api_key', 'VARCHAR', 200, 'N', 'Y', '', 'API 인증키'),
                 ('요청엔드포인트', 'endpoint', 'VARCHAR', 500, 'N', 'N', '', '/api/v2/...'),
                 ('HTTP메서드', 'http_method', 'CHAR', 6, 'N', 'N', '', 'GET/POST/PUT/DELETE'),
                 ('요청IP', 'client_ip', 'VARCHAR', 45, 'N', 'N', '', 'IPv4/IPv6'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_SYS_API_LOG', 'k_dir': '수신',
             'fields': [
                 ('요청ID', 'request_id', 'VARCHAR', 36, 'N', 'Y', '', 'UUID, PK'),
                 ('응답코드', 'status_code', 'CHAR', 3, 'N', 'N', '', 'HTTP 상태코드'),
                 ('응답시간', 'latency_ms', 'NUMERIC', 10, 'N', 'N', '', '단위: ms'),
                 ('응답크기', 'response_size', 'NUMERIC', 10, 'N', 'N', '', '단위: bytes'),
                 ('인증결과', 'auth_result', 'CHAR', 1, 'N', 'N', '', 'Y/N'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '/api/v2/*', 'k_dir': '수신',
             'fields': [
                 ('Rate Limit잔여', 'rate_remaining', 'NUMERIC', 10, 'Y', 'N', '', '분당 허용 잔여 수'),
                 ('결과메시지', 'result_msg', 'VARCHAR', 200, 'N', 'N', '', ''),
             ]},
        ]
    },
]

# ============================================================
# Batch 2: 데이터수집 - 외부 (5건)
# ============================================================
batch_co_ext = [
    {
        'area': '데이터수집', 'if_id': 'IF-CO-001',
        'if_name': '기상청 날씨 API 수집',
        'req_sys': '데이터허브포털', 'res_sys': '기상청 Open API',
        'in_out': '외부', 'cycle': '매 1시간', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '외부연계 현황', 'k_dir': '송신',
             'fields': [
                 ('인증키', 'service_key', 'VARCHAR', 200, 'N', 'Y', '', '공공데이터포털 발급키'),
                 ('기준일자', 'base_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
                 ('기준시각', 'base_time', 'CHAR', 4, 'N', 'N', '', 'HHmm'),
                 ('X좌표', 'nx', 'NUMERIC', 3, 'N', 'N', '', '격자 X좌표'),
                 ('Y좌표', 'ny', 'NUMERIC', 3, 'N', 'N', '', '격자 Y좌표'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_WEATHER', 'k_dir': '수신',
             'fields': [
                 ('결과코드', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                 ('관측일시', 'obs_date', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('기온', 'temperature', 'NUMERIC', 5, 'Y', 'N', '', '단위: ℃'),
                 ('습도', 'humidity', 'NUMERIC', 5, 'Y', 'N', '', '단위: %'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'https://apis.data.go.kr/1360000/VilageFcstInfoService', 'k_dir': '수신',
             'fields': [
                 ('강수량', 'precipitation', 'NUMERIC', 8, 'Y', 'N', '', '단위: mm'),
                 ('풍속', 'wind_speed', 'NUMERIC', 5, 'Y', 'N', '', '단위: m/s'),
                 ('결과메시지', 'result_msg', 'VARCHAR', 200, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-002',
        'if_name': '환경부 수질측정 데이터 수집',
        'req_sys': '데이터허브포털', 'res_sys': '환경부 수질정보시스템',
        'in_out': '외부', 'cycle': '매 3시간', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '외부연계 현황', 'k_dir': '송신',
             'fields': [
                 ('인증키', 'service_key', 'VARCHAR', 200, 'N', 'Y', '', '공공데이터포털 발급키'),
                 ('측정소코드', 'site_id', 'VARCHAR', 20, 'N', 'N', '', ''),
                 ('조회시작일', 'start_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
                 ('조회종료일', 'end_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_WATER_QUALITY', 'k_dir': '수신',
             'fields': [
                 ('결과코드', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                 ('측정소명', 'site_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('측정일시', 'meas_date', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('수소이온농도', 'ph', 'NUMERIC', 5, 'Y', 'N', '', '단위: pH'),
                 ('용존산소', 'do_value', 'NUMERIC', 5, 'Y', 'N', '', '단위: mg/L'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'https://apis.data.go.kr/B500001/rwis/waterQuality', 'k_dir': '수신',
             'fields': [
                 ('BOD', 'bod', 'NUMERIC', 5, 'Y', 'N', '', '단위: mg/L'),
                 ('탁도', 'turbidity', 'NUMERIC', 8, 'Y', 'N', '', '단위: NTU'),
                 ('잔류염소', 'residual_cl', 'NUMERIC', 5, 'Y', 'N', '', '단위: mg/L'),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-003',
        'if_name': '수자원공사 댐수위 관측 데이터 수집',
        'req_sys': '데이터허브포털', 'res_sys': '수자원공사 SFTP서버',
        'in_out': '외부', 'cycle': '매 6시간', 'method': 'SFTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '외부연계 현황', 'k_dir': '송신',
             'fields': [
                 ('SFTP호스트', 'sftp_host', 'VARCHAR', 100, 'N', 'N', '', '접속 호스트'),
                 ('SFTP포트', 'sftp_port', 'NUMERIC', 5, 'N', 'N', '22', ''),
                 ('접속계정', 'sftp_user', 'VARCHAR', 50, 'N', 'N', '', ''),
                 ('원격경로', 'remote_path', 'VARCHAR', 500, 'N', 'N', '', '파일 수신 디렉토리'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_DAM_LEVEL', 'k_dir': '수신',
             'fields': [
                 ('댐코드', 'dam_code', 'VARCHAR', 10, 'N', 'Y', '', 'PK'),
                 ('관측일시', 'obs_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                 ('수위', 'water_level', 'NUMERIC', 8, 'N', 'N', '', '단위: EL.m'),
                 ('유입량', 'inflow', 'NUMERIC', 10, 'Y', 'N', '', '단위: m³/s'),
                 ('방류량', 'outflow', 'NUMERIC', 10, 'Y', 'N', '', '단위: m³/s'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'sftp://data.kwater.or.kr/dam/', 'k_dir': '수신',
             'fields': [
                 ('저수율', 'storage_rate', 'NUMERIC', 5, 'Y', 'N', '', '단위: %'),
                 ('파일명', 'file_name', 'VARCHAR', 200, 'N', 'N', '', 'CSV 파일'),
                 ('수신건수', 'recv_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-004',
        'if_name': '한국환경공단 환경영향평가 데이터 수집',
        'req_sys': '데이터허브포털', 'res_sys': '한국환경공단 API',
        'in_out': '외부', 'cycle': '일 1회', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '외부연계 현황', 'k_dir': '송신',
             'fields': [
                 ('인증키', 'service_key', 'VARCHAR', 200, 'N', 'Y', '', '공공데이터포털 발급키'),
                 ('평가유형', 'eval_type', 'VARCHAR', 20, 'N', 'N', '', '환경영향/소규모'),
                 ('조회기간시작', 'from_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
                 ('조회기간종료', 'to_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_ENV_ASSESS', 'k_dir': '수신',
             'fields': [
                 ('결과코드', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                 ('평가번호', 'assess_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                 ('사업명', 'project_name', 'VARCHAR', 200, 'N', 'N', '', ''),
                 ('평가등급', 'assess_grade', 'CHAR', 1, 'Y', 'N', '', 'A/B/C/D'),
                 ('평가일자', 'assess_date', 'DATETIME', 23, 'N', 'N', '', ''),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'https://apis.data.go.kr/B553006/EIAService', 'k_dir': '수신',
             'fields': [
                 ('평가내용요약', 'summary', 'VARCHAR', 2000, 'Y', 'N', '', ''),
                 ('결과메시지', 'result_msg', 'VARCHAR', 200, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-005',
        'if_name': '기상위성 영상 데이터 수집',
        'req_sys': '데이터허브포털', 'res_sys': '기상청 위성센터',
        'in_out': '외부', 'cycle': '일 1회', 'method': 'SFTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '외부연계 현황', 'k_dir': '송신',
             'fields': [
                 ('SFTP호스트', 'sftp_host', 'VARCHAR', 100, 'N', 'N', '', '접속 호스트'),
                 ('접속계정', 'sftp_user', 'VARCHAR', 50, 'N', 'N', '', ''),
                 ('원격경로', 'remote_path', 'VARCHAR', 500, 'N', 'N', '', '영상 디렉토리'),
                 ('영상유형', 'image_type', 'VARCHAR', 20, 'N', 'N', '', 'IR/VIS/WV'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_SATELLITE_IMG', 'k_dir': '수신',
             'fields': [
                 ('영상ID', 'image_id', 'VARCHAR', 50, 'N', 'Y', '', 'PK'),
                 ('촬영일시', 'capture_datetime', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('영상유형', 'image_type', 'VARCHAR', 20, 'N', 'N', '', 'IR/VIS/WV'),
                 ('파일경로', 'file_path', 'VARCHAR', 500, 'N', 'N', '', '저장경로'),
                 ('파일크기', 'file_size', 'NUMERIC', 15, 'N', 'N', '', '단위: bytes'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'sftp://satellite.kma.go.kr/image/', 'k_dir': '수신',
             'fields': [
                 ('해상도', 'resolution', 'VARCHAR', 20, 'Y', 'N', '', '단위: km'),
                 ('수신건수', 'recv_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
]

# ============================================================
# Batch 3: 데이터수집 - 내부 (5건)
# ============================================================
batch_co_int = [
    {
        'area': '데이터수집', 'if_id': 'IF-CO-006',
        'if_name': 'SAP HANA 자산 CDC 연계',
        'req_sys': 'SAP HANA DB', 'res_sys': '데이터허브포털',
        'in_out': '내부', 'cycle': '실시간', 'method': 'CDC',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'CDC 연계현황', 'k_dir': '송신',
             'fields': [
                 ('CDC커넥터', 'connector_name', 'VARCHAR', 50, 'N', 'N', '', 'cdc-sap-asset'),
                 ('원천DB', 'source_db', 'VARCHAR', 50, 'N', 'N', '', 'SAP HANA'),
                 ('대상테이블수', 'table_count', 'NUMERIC', 5, 'N', 'N', '', ''),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_SAP_ASSET', 'k_dir': '수신',
             'fields': [
                 ('자산번호', 'asset_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                 ('자산명', 'asset_name', 'VARCHAR', 200, 'N', 'N', '', ''),
                 ('취득일자', 'acq_date', 'DATETIME', 23, 'Y', 'N', '', ''),
                 ('취득금액', 'acq_amount', 'NUMERIC', 15, 'Y', 'N', '', '단위: 원'),
                 ('자산상태', 'asset_status', 'CHAR', 1, 'N', 'N', '', 'A/D/R'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'kafka://datahub-broker:9092/cdc_sap_asset', 'k_dir': '수신',
             'fields': [
                 ('이벤트유형', 'op_type', 'CHAR', 1, 'N', 'N', '', 'I/U/D'),
                 ('이벤트일시', 'event_ts', 'DATETIME', 23, 'N', 'N', '', 'CDC 타임스탬프'),
                 ('처리건수', 'events_count', 'NUMERIC', 10, 'N', 'N', '', 'Events/분'),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-007',
        'if_name': 'Oracle 수질DB CDC 연계',
        'req_sys': 'Oracle DB', 'res_sys': '데이터허브포털',
        'in_out': '내부', 'cycle': '실시간', 'method': 'CDC',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'CDC 연계현황', 'k_dir': '송신',
             'fields': [
                 ('CDC커넥터', 'connector_name', 'VARCHAR', 50, 'N', 'N', '', 'cdc-ora-water'),
                 ('원천DB', 'source_db', 'VARCHAR', 50, 'N', 'N', '', 'Oracle'),
                 ('대상테이블수', 'table_count', 'NUMERIC', 5, 'N', 'N', '', ''),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_ORA_WATER_QUALITY', 'k_dir': '수신',
             'fields': [
                 ('측정ID', 'meas_id', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                 ('정수장코드', 'plant_code', 'VARCHAR', 10, 'N', 'N', '', ''),
                 ('측정일시', 'meas_datetime', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('pH', 'ph_value', 'NUMERIC', 5, 'Y', 'N', '', ''),
                 ('잔류염소', 'cl_residual', 'NUMERIC', 5, 'Y', 'N', '', '단위: mg/L'),
                 ('탁도', 'turbidity', 'NUMERIC', 8, 'Y', 'N', '', '단위: NTU'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'kafka://datahub-broker:9092/cdc_ora_water', 'k_dir': '수신',
             'fields': [
                 ('이벤트유형', 'op_type', 'CHAR', 1, 'N', 'N', '', 'I/U/D'),
                 ('이벤트일시', 'event_ts', 'DATETIME', 23, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-008',
        'if_name': '티베로 수위관측DB 연계',
        'req_sys': '티베로 DB (FA망)', 'res_sys': '데이터허브포털',
        'in_out': '내부', 'cycle': '매 30분', 'method': 'DB Link',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '수집현황 모니터링', 'k_dir': '송신',
             'fields': [
                 ('원천DB', 'source_db', 'VARCHAR', 50, 'N', 'N', '', '티베로/계측DB'),
                 ('구간', 'zone', 'VARCHAR', 10, 'N', 'N', '', 'FA'),
                 ('수집방식', 'collect_method', 'VARCHAR', 10, 'N', 'N', '', '배치'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_RWIS_LEVEL', 'k_dir': '수신',
             'fields': [
                 ('관측소코드', 'station_code', 'VARCHAR', 10, 'N', 'Y', '', 'PK'),
                 ('관측일시', 'obs_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                 ('수위', 'water_level', 'NUMERIC', 8, 'N', 'N', '', '단위: m'),
                 ('유량', 'flow_rate', 'NUMERIC', 10, 'Y', 'N', '', '단위: m³/s'),
                 ('수온', 'water_temp', 'NUMERIC', 5, 'Y', 'N', '', '단위: ℃'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'jdbc:tibero://fa-rwis-db:8629/RWISDB', 'k_dir': '수신',
             'fields': [
                 ('처리건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
                 ('최종동기화', 'last_sync', 'DATETIME', 23, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-009',
        'if_name': 'arcGIS 공간정보 연계',
        'req_sys': 'arcGIS Server', 'res_sys': '데이터허브포털',
        'in_out': '내부', 'cycle': '일 1회', 'method': 'DB Link',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '대내외시스템 연계관리', 'k_dir': '송신',
             'fields': [
                 ('원천시스템', 'source_system', 'VARCHAR', 50, 'N', 'N', '', 'arcGIS'),
                 ('연계방식', 'link_method', 'VARCHAR', 20, 'N', 'N', '', 'DB Link'),
                 ('대상레이어', 'target_layer', 'VARCHAR', 100, 'N', 'N', '', '공간정보 레이어'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_GIS_FACILITY', 'k_dir': '수신',
             'fields': [
                 ('시설물ID', 'facility_id', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                 ('시설물유형', 'facility_type', 'VARCHAR', 20, 'N', 'N', '', '관로/밸브/계량기'),
                 ('위도', 'latitude', 'NUMERIC', 12, 'N', 'N', '', 'WGS84'),
                 ('경도', 'longitude', 'NUMERIC', 12, 'N', 'N', '', 'WGS84'),
                 ('GIS좌표', 'geom_wkt', 'TEXT', '', 'Y', 'N', '', 'WKT 형식'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'jdbc:oracle://gis-db:1521/GISDB', 'k_dir': '수신',
             'fields': [
                 ('좌표계', 'srid', 'NUMERIC', 6, 'N', 'N', '4326', 'EPSG 코드'),
                 ('레코드수', 'record_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터수집', 'if_id': 'IF-CO-010',
        'if_name': 'IoT 센서 실시간 스트리밍 수집',
        'req_sys': 'IoT 센서게이트웨이', 'res_sys': '데이터허브포털',
        'in_out': '내부', 'cycle': '실시간', 'method': 'Kafka',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'Kafka 스트리밍', 'k_dir': '송신',
             'fields': [
                 ('토픽명', 'topic_name', 'VARCHAR', 100, 'N', 'N', '', 'iot_sensor_raw'),
                 ('파티션수', 'partitions', 'NUMERIC', 3, 'N', 'N', '6', ''),
                 ('복제팩터', 'replication', 'NUMERIC', 1, 'N', 'N', '3', ''),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_COL_IOT_SENSOR', 'k_dir': '수신',
             'fields': [
                 ('센서ID', 'sensor_id', 'VARCHAR', 30, 'N', 'Y', '', 'PK'),
                 ('측정일시', 'meas_ts', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                 ('센서유형', 'sensor_type', 'VARCHAR', 20, 'N', 'N', '', '수위/수압/유량/수질'),
                 ('측정값', 'meas_value', 'NUMERIC', 15, 'N', 'N', '', ''),
                 ('단위', 'unit', 'VARCHAR', 10, 'Y', 'N', '', 'm/kPa/m³등'),
                 ('배터리잔량', 'battery_pct', 'NUMERIC', 3, 'Y', 'N', '', '단위: %'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'kafka://datahub-broker:9092/iot_sensor_raw', 'k_dir': '수신',
             'fields': [
                 ('메시지지연', 'msg_lag', 'NUMERIC', 10, 'Y', 'N', '', '단위: ms'),
                 ('수신율', 'in_rate', 'NUMERIC', 10, 'N', 'N', '', 'msg/s'),
             ]},
        ]
    },
]

# ============================================================
# Batch 4: 데이터배포 (5건) + 메타데이터 (2건)
# ============================================================
batch_di = [
    {
        'area': '데이터배포', 'if_id': 'IF-DI-001',
        'if_name': '광역상수도 유량 실시간 API',
        'req_sys': '외부기관/시스템', 'res_sys': '데이터허브포털',
        'in_out': '외부', 'cycle': '실시간', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'Data Product 목록', 'k_dir': '송신',
             'fields': [
                 ('API인증키', 'api_key', 'VARCHAR', 200, 'N', 'Y', '', ''),
                 ('지역코드', 'region_code', 'VARCHAR', 10, 'Y', 'N', '', ''),
                 ('시설코드', 'facility_id', 'VARCHAR', 20, 'Y', 'N', '', ''),
                 ('조회시작일시', 'from_date', 'DATETIME', 23, 'N', 'N', '', 'from'),
                 ('조회종료일시', 'to_date', 'DATETIME', 23, 'N', 'N', '', 'to'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_DIST_FLOW_REALTIME', 'k_dir': '수신',
             'fields': [
                 ('응답코드', 'status_code', 'CHAR', 3, 'N', 'N', '', 'HTTP 상태코드'),
                 ('총건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
                 ('시설명', 'facility_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('측정일시', 'meas_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '/api/v2/water/flow/realtime', 'k_dir': '수신',
             'fields': [
                 ('유량', 'flow_rate', 'NUMERIC', 10, 'N', 'N', '', '단위: m³/s'),
                 ('수압', 'pressure', 'NUMERIC', 8, 'Y', 'N', '', '단위: kPa'),
                 ('수위', 'water_level', 'NUMERIC', 8, 'Y', 'N', '', '단위: m'),
             ]},
        ]
    },
    {
        'area': '데이터배포', 'if_id': 'IF-DI-002',
        'if_name': '수력발전 일간통계 API',
        'req_sys': '외부기관/시스템', 'res_sys': '데이터허브포털',
        'in_out': '외부', 'cycle': '일 1회', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'Data Product 목록', 'k_dir': '송신',
             'fields': [
                 ('API인증키', 'api_key', 'VARCHAR', 200, 'N', 'Y', '', ''),
                 ('발전소코드', 'plant_code', 'VARCHAR', 10, 'Y', 'N', '', ''),
                 ('통계일자', 'stat_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_DIST_POWER_DAILY', 'k_dir': '수신',
             'fields': [
                 ('응답코드', 'status_code', 'CHAR', 3, 'N', 'N', '', ''),
                 ('발전소명', 'plant_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('통계일자', 'stat_date', 'CHAR', 8, 'N', 'Y', '', 'PK'),
                 ('발전량', 'power_gen', 'NUMERIC', 12, 'N', 'N', '', '단위: MWh'),
                 ('가동시간', 'oper_hours', 'NUMERIC', 5, 'Y', 'N', '', '단위: 시간'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '/api/v2/energy/power/daily', 'k_dir': '수신',
             'fields': [
                 ('발전효율', 'efficiency', 'NUMERIC', 5, 'Y', 'N', '', '단위: %'),
                 ('총건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터배포', 'if_id': 'IF-DI-003',
        'if_name': '수질측정 데이터 API',
        'req_sys': '외부기관/시스템', 'res_sys': '데이터허브포털',
        'in_out': '외부', 'cycle': '매 30분', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'Data Product 목록', 'k_dir': '송신',
             'fields': [
                 ('API인증키', 'api_key', 'VARCHAR', 200, 'N', 'Y', '', ''),
                 ('정수장코드', 'plant_code', 'VARCHAR', 10, 'Y', 'N', '', ''),
                 ('측정항목', 'measure_item', 'VARCHAR', 20, 'Y', 'N', '', 'pH/DO/탁도 등'),
                 ('조회시작일', 'from_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_DIST_WATER_QUALITY', 'k_dir': '수신',
             'fields': [
                 ('응답코드', 'status_code', 'CHAR', 3, 'N', 'N', '', ''),
                 ('측정소명', 'site_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('측정일시', 'meas_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                 ('pH', 'ph', 'NUMERIC', 5, 'Y', 'N', '', ''),
                 ('용존산소', 'do_value', 'NUMERIC', 5, 'Y', 'N', '', '단위: mg/L'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '/api/v2/water/quality', 'k_dir': '수신',
             'fields': [
                 ('탁도', 'turbidity', 'NUMERIC', 8, 'Y', 'N', '', '단위: NTU'),
                 ('잔류염소', 'residual_cl', 'NUMERIC', 5, 'Y', 'N', '', '단위: mg/L'),
                 ('총건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터배포', 'if_id': 'IF-DI-004',
        'if_name': '댐수위·방류량 종합 API',
        'req_sys': '외부기관/시스템', 'res_sys': '데이터허브포털',
        'in_out': '외부', 'cycle': '실시간', 'method': 'HTTP',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'Data Product 목록', 'k_dir': '송신',
             'fields': [
                 ('API인증키', 'api_key', 'VARCHAR', 200, 'N', 'Y', '', ''),
                 ('댐코드', 'dam_code', 'VARCHAR', 10, 'Y', 'N', '', ''),
                 ('조회시작일시', 'from_date', 'DATETIME', 23, 'N', 'N', '', 'from'),
                 ('조회종료일시', 'to_date', 'DATETIME', 23, 'N', 'N', '', 'to'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_DIST_DAM_COMPOSITE', 'k_dir': '수신',
             'fields': [
                 ('응답코드', 'status_code', 'CHAR', 3, 'N', 'N', '', ''),
                 ('댐명', 'dam_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('관측일시', 'obs_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                 ('수위', 'water_level', 'NUMERIC', 8, 'N', 'N', '', '단위: EL.m'),
                 ('유입량', 'inflow', 'NUMERIC', 10, 'Y', 'N', '', '단위: m³/s'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '/api/v2/dam/composite', 'k_dir': '수신',
             'fields': [
                 ('방류량', 'outflow', 'NUMERIC', 10, 'Y', 'N', '', '단위: m³/s'),
                 ('저수율', 'storage_rate', 'NUMERIC', 5, 'Y', 'N', '', '단위: %'),
                 ('총건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '데이터배포', 'if_id': 'IF-DI-005',
        'if_name': '원격검침 사용량분석 스트리밍',
        'req_sys': '데이터허브포털', 'res_sys': '수요분석시스템',
        'in_out': '내부', 'cycle': '실시간', 'method': 'Kafka',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'Data Product 목록', 'k_dir': '송신',
             'fields': [
                 ('토픽명', 'topic_name', 'VARCHAR', 100, 'N', 'N', '', 'smart_meter_read'),
                 ('파티션키', 'partition_key', 'VARCHAR', 50, 'N', 'N', '', '고객번호 기준'),
                 ('보존기간', 'retention', 'VARCHAR', 10, 'N', 'N', '7d', ''),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_DIST_METER_USAGE', 'k_dir': '수신',
             'fields': [
                 ('고객번호', 'customer_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                 ('검침일시', 'read_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                 ('사용량', 'usage_amount', 'NUMERIC', 12, 'N', 'N', '', '단위: m³'),
                 ('미터ID', 'meter_id', 'VARCHAR', 20, 'N', 'N', '', ''),
                 ('이상여부', 'anomaly_yn', 'CHAR', 1, 'N', 'N', 'N', 'Y/N'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': 'kafka://datahub-broker:9092/smart_meter_read', 'k_dir': '수신',
             'fields': [
                 ('메시지율', 'out_rate', 'NUMERIC', 10, 'N', 'N', '', 'msg/s'),
                 ('소비자그룹', 'consumer_group', 'VARCHAR', 50, 'N', 'N', '', ''),
             ]},
        ]
    },
]

batch_me = [
    {
        'area': '메타데이터', 'if_id': 'IF-ME-001',
        'if_name': '메타데이터 카탈로그 동기화',
        'req_sys': '데이터허브포털', 'res_sys': '원천시스템 DB',
        'in_out': '내부', 'cycle': '일 1회', 'method': 'DB',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': '데이터모델 관리', 'k_dir': '송신',
             'fields': [
                 ('동기화대상', 'sync_target', 'VARCHAR', 20, 'N', 'N', '', '테이블/컬럼/관계'),
                 ('원천DB유형', 'db_type', 'VARCHAR', 20, 'N', 'N', '', 'SAP/Oracle/Tibero'),
                 ('스키마명', 'schema_name', 'VARCHAR', 50, 'N', 'N', '', ''),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_META_CATALOG', 'k_dir': '수신',
             'fields': [
                 ('테이블ID', 'table_id', 'VARCHAR', 50, 'N', 'Y', '', 'PK'),
                 ('테이블명', 'table_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('테이블설명', 'table_desc', 'VARCHAR', 500, 'Y', 'N', '', ''),
                 ('컬럼수', 'column_count', 'NUMERIC', 5, 'N', 'N', '', ''),
                 ('PII컬럼수', 'pii_count', 'NUMERIC', 5, 'N', 'N', '0', '개인정보 포함 컬럼'),
                 ('도메인', 'domain', 'VARCHAR', 20, 'N', 'N', '', '수자원/에너지/고객 등'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '-', 'k_dir': '수신',
             'fields': [
                 ('표준준수율', 'compliance_rate', 'NUMERIC', 5, 'N', 'N', '', '단위: %'),
                 ('최종수정일', 'last_modified', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('처리건수', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
             ]},
        ]
    },
    {
        'area': '메타데이터', 'if_id': 'IF-ME-002',
        'if_name': '데이터 품질(DQ) 검증 결과 연계',
        'req_sys': 'DQ엔진', 'res_sys': '데이터허브포털',
        'in_out': '내부', 'cycle': '일 1회', 'method': 'DB',
        'sections': [
            {'i_type': '관련\n화면', 'j_name': 'DQ 검증·프로파일링', 'k_dir': '송신',
             'fields': [
                 ('검증규칙ID', 'rule_id', 'VARCHAR', 20, 'N', 'Y', '', ''),
                 ('대상테이블', 'target_table', 'VARCHAR', 100, 'N', 'N', '', ''),
                 ('대상필드', 'target_field', 'VARCHAR', 100, 'Y', 'N', '', ''),
                 ('검증유형', 'check_type', 'VARCHAR', 20, 'N', 'N', '', '완전성/유효성/정합성'),
             ]},
            {'i_type': '관련\n테이블', 'j_name': 'TB_META_DQ_RESULT', 'k_dir': '수신',
             'fields': [
                 ('검증ID', 'check_id', 'VARCHAR', 36, 'N', 'Y', '', 'UUID, PK'),
                 ('실행일시', 'exec_datetime', 'DATETIME', 23, 'N', 'N', '', ''),
                 ('결과', 'result', 'VARCHAR', 10, 'N', 'N', '', '통과/경고/실패'),
                 ('위반건수', 'violation_count', 'NUMERIC', 10, 'N', 'N', '0', ''),
                 ('임계값', 'threshold', 'NUMERIC', 5, 'N', 'N', '', '단위: %'),
             ]},
            {'i_type': '관련\nAPI/URL', 'j_name': '-', 'k_dir': '수신',
             'fields': [
                 ('추이', 'trend', 'CHAR', 1, 'Y', 'N', '', '▲/▼/—'),
                 ('담당자', 'owner', 'VARCHAR', 50, 'Y', 'N', '', ''),
             ]},
        ]
    },
]

# ============================================================
# 전체 합치기 및 엑셀 생성
# ============================================================
all_interfaces = batch_sy + batch_co_ext + batch_co_int + batch_di + batch_me

row = 4
for iface in all_interfaces:
    row = write_interface(ws, row, iface)

# 표지 업데이트
ws_cover = wb.worksheets[0]
ws_cover['C17'] = 'KWDP-DH-DS-17'
ws_cover['C19'] = '0.1'

wb.save(DST)
print(f'저장 완료: {DST}')
print(f'총 인터페이스: {len(all_interfaces)}건')
print(f'총 {row - 4}행 (row 4 ~ row {row - 1})')
print()
for iface in all_interfaces:
    fields = sum(len(s['fields']) for s in iface['sections'])
    print(f"  {iface['if_id']}: {iface['if_name']} ({fields}필드)")
