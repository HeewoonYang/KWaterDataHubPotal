# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = r'D:\00_수공프로젝트\20260224_설계단계_산출물\산출물 템플릿_데이터허브(손이사님 제공)\4. 데이터허브\2. 설계\(템플릿)KWDP-DH-DS-17-인터페이스설계서-v0.x.xlsx'
dst = r'D:\00_수공프로젝트\20260224_설계단계_산출물\산출물 템플릿_데이터허브(손이사님 제공)\4. 데이터허브\2. 설계\KWDP-DH-DS-17-인터페이스설계서-v0.1.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb.worksheets[2]

thin = Side(style='thin')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
font_data = Font(name='\ub9d1\uc740 \uace0\ub515', size=10)
font_bold = Font(name='\ub9d1\uc740 \uace0\ub515', size=10, bold=True)
fill_gray = PatternFill('solid', fgColor='F2F2F2')
align_cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_lc = Alignment(horizontal='left', vertical='center', wrap_text=True)


def set_cell(ws, row, col, value, bold=False, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_bold if bold else font_data
    c.border = border_all
    c.alignment = align if align else align_cc
    if fill:
        c.fill = fill
    return c


def set_empty_border(ws, row, col, align=None):
    c = ws.cell(row=row, column=col)
    c.font = font_data
    c.border = border_all
    c.alignment = align if align else align_cc


def write_interface(ws, start_row, data):
    total_rows = sum(len(s['fields']) for s in data['sections'])
    end_row = start_row + total_rows - 1

    basic_cols = [
        (1, data['area']), (2, data['if_id']), (3, data['if_name']),
        (4, data['req_sys']), (5, data['res_sys']), (6, data['in_out']),
        (7, data['cycle']), (8, data['method'])
    ]
    for col, val in basic_cols:
        al = align_lc if col == 3 else align_cc
        set_cell(ws, start_row, col, val, align=al)
        if total_rows > 1:
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row, end_column=col)
            for r in range(start_row + 1, end_row + 1):
                set_empty_border(ws, r, col, al)

    cur_row = start_row
    k_groups = []
    cur_k_dir = None
    cur_k_start = None

    for sec in data['sections']:
        sec_start = cur_row
        sec_end = cur_row + len(sec['fields']) - 1

        set_cell(ws, sec_start, 9, sec['i_type'], bold=True, fill=fill_gray)
        if len(sec['fields']) > 1:
            ws.merge_cells(start_row=sec_start, start_column=9,
                           end_row=sec_end, end_column=9)
            for r in range(sec_start + 1, sec_end + 1):
                set_empty_border(ws, r, 9)

        set_cell(ws, sec_start, 10, sec['j_name'], align=align_lc)
        if len(sec['fields']) > 1:
            ws.merge_cells(start_row=sec_start, start_column=10,
                           end_row=sec_end, end_column=10)
            for r in range(sec_start + 1, sec_end + 1):
                set_empty_border(ws, r, 10, align_lc)

        actual_k = sec['k_dir']
        for i, field in enumerate(sec['fields']):
            r = cur_row + i
            if cur_k_dir is None:
                cur_k_dir = actual_k
                cur_k_start = r
            elif actual_k != cur_k_dir:
                k_groups.append((cur_k_start, r - 1, cur_k_dir))
                cur_k_dir = actual_k
                cur_k_start = r

            set_cell(ws, r, 12, field[0], align=align_lc)
            set_cell(ws, r, 13, field[1], align=align_lc)
            set_cell(ws, r, 14, field[2], align=align_lc)
            set_cell(ws, r, 15, field[3])
            set_cell(ws, r, 16, field[4])
            set_cell(ws, r, 17, field[5])
            set_cell(ws, r, 18, field[6] if len(field) > 6 else '', align=align_lc)
            set_cell(ws, r, 19, field[7] if len(field) > 7 else '', align=align_lc)

        cur_row = sec_end + 1

    if cur_k_start is not None:
        k_groups.append((cur_k_start, end_row, cur_k_dir))

    for ks, ke, kd in k_groups:
        set_cell(ws, ks, 11, kd, bold=True, fill=fill_gray)
        if ke > ks:
            ws.merge_cells(start_row=ks, start_column=11,
                           end_row=ke, end_column=11)
            for r in range(ks + 1, ke + 1):
                set_empty_border(ws, r, 11)

    return end_row + 1


interfaces = [
    {
        'area': '\uc2dc\uc2a4\ud15c\uad00\ub9ac',
        'if_id': 'IF-DH-001',
        'if_name': 'ERP \uc778\uc0ac\uc815\ubcf4 \ub3d9\uae30\ud654',
        'req_sys': '\ub370\uc774\ud130\ud5c8\ube0c\ud3ec\ud138',
        'res_sys': 'K-water ERP(SAP)',
        'in_out': '\ub0b4\ubd80',
        'cycle': '\uc77c 1\ud68c',
        'method': 'DB',
        'sections': [
            {
                'i_type': '\uad00\ub828\n\ud654\uba74',
                'j_name': 'ERP \uc778\uc0ac\uc815\ubcf4 \ub3d9\uae30\ud654',
                'k_dir': '\uc1a1\uc2e0',
                'fields': [
                    ('\ub3d9\uae30\ud654\uc720\ud615', 'sync_type', 'VARCHAR', 10, 'N', 'N', '', '\uc804\uccb4/\uc99d\ubd84/CDC'),
                    ('\ub300\uc0c1\uad6c\ubd84', 'target_type', 'VARCHAR', 10, 'N', 'N', '', '\uc778\uc0ac/\uc870\uc9c1/\uacb8\uc9c1'),
                    ('\uc694\uccad\uc77c\uc2dc', 'req_date', 'DATETIME', 23, 'N', 'N', '', ''),
                ]
            },
            {
                'i_type': '\uad00\ub828\n\ud14c\uc774\ube14',
                'j_name': 'TB_SYS_USER',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uc0ac\ubc88', 'emp_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK, PREFIX("KW")'),
                    ('\uc774\ub984', 'user_name', 'VARCHAR', 50, 'N', 'N', '', ''),
                    ('\ubd80\uc11c\ucf54\ub4dc', 'dept_code', 'VARCHAR', 20, 'N', 'N', '', 'LOOKUP(ORG_MAP)'),
                    ('\uc9c1\uae09', 'position', 'VARCHAR', 20, 'Y', 'N', '', 'LOOKUP(JOB_MAP)'),
                    ('\uc774\uba54\uc77c', 'email', 'VARCHAR', 100, 'Y', 'N', '', 'LOWER()'),
                ]
            },
            {
                'i_type': '\uad00\ub828\nAPI/URL',
                'j_name': '-',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uc0c1\ud0dc', 'status', 'CHAR', 1, 'N', 'N', 'A', 'MAP(A\u2192\ud65c\uc131,T\u2192\ud1f4\uc9c1)'),
                    ('\ucc98\ub9ac\uac74\uc218', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
                    ('\uacb0\uacfc\ucf54\ub4dc', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                ]
            },
        ]
    },
    {
        'area': '\ub370\uc774\ud130\uc218\uc9d1',
        'if_id': 'IF-DH-002',
        'if_name': '\uae30\uc0c1\uccad \ub0a0\uc528 API \uc218\uc9d1',
        'req_sys': '\ub370\uc774\ud130\ud5c8\ube0c\ud3ec\ud138',
        'res_sys': '\uae30\uc0c1\uccad Open API',
        'in_out': '\uc678\ubd80',
        'cycle': '\ub9e4 1\uc2dc\uac04',
        'method': 'HTTP',
        'sections': [
            {
                'i_type': '\uad00\ub828\n\ud654\uba74',
                'j_name': '\uc678\ubd80\uc5f0\uacc4 \ud604\ud669',
                'k_dir': '\uc1a1\uc2e0',
                'fields': [
                    ('\uc778\uc99d\ud0a4', 'service_key', 'VARCHAR', 200, 'N', 'Y', '', '\uacf5\uacf5\ub370\uc774\ud130\ud3ec\ud138 \ubc1c\uae09\ud0a4'),
                    ('\uae30\uc900\uc77c\uc790', 'base_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
                    ('\uae30\uc900\uc2dc\uac01', 'base_time', 'CHAR', 4, 'N', 'N', '', 'HHmm'),
                    ('X\uc88c\ud45c', 'nx', 'NUMERIC', 3, 'N', 'N', '', '\uaca9\uc790 X\uc88c\ud45c'),
                    ('Y\uc88c\ud45c', 'ny', 'NUMERIC', 3, 'N', 'N', '', '\uaca9\uc790 Y\uc88c\ud45c'),
                ]
            },
            {
                'i_type': '\uad00\ub828\n\ud14c\uc774\ube14',
                'j_name': 'TB_COL_WEATHER',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uacb0\uacfc\ucf54\ub4dc', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                    ('\uad00\uce21\uc77c\uc2dc', 'obs_date', 'DATETIME', 23, 'N', 'N', '', ''),
                    ('\uae30\uc628', 'temperature', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: \u2103'),
                    ('\uc2b5\ub3c4', 'humidity', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: %'),
                ]
            },
            {
                'i_type': '\uad00\ub828\nAPI/URL',
                'j_name': 'https://apis.data.go.kr/1360000/VilageFcstInfoService',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uac15\uc218\ub7c9', 'precipitation', 'NUMERIC', 8, 'Y', 'N', '', '\ub2e8\uc704: mm'),
                    ('\ud48d\uc18d', 'wind_speed', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: m/s'),
                    ('\uacb0\uacfc\uba54\uc2dc\uc9c0', 'result_msg', 'VARCHAR', 200, 'N', 'N', '', ''),
                ]
            },
        ]
    },
    {
        'area': '\ub370\uc774\ud130\uc218\uc9d1',
        'if_id': 'IF-DH-003',
        'if_name': '\ud658\uacbd\ubd80 \uc218\uc9c8\uce21\uc815 \ub370\uc774\ud130 \uc218\uc9d1',
        'req_sys': '\ub370\uc774\ud130\ud5c8\ube0c\ud3ec\ud138',
        'res_sys': '\ud658\uacbd\ubd80 \uc218\uc9c8\uc815\ubcf4\uc2dc\uc2a4\ud15c',
        'in_out': '\uc678\ubd80',
        'cycle': '\ub9e4 3\uc2dc\uac04',
        'method': 'HTTP',
        'sections': [
            {
                'i_type': '\uad00\ub828\n\ud654\uba74',
                'j_name': '\uc678\ubd80\uc5f0\uacc4 \ud604\ud669',
                'k_dir': '\uc1a1\uc2e0',
                'fields': [
                    ('\uc778\uc99d\ud0a4', 'service_key', 'VARCHAR', 200, 'N', 'Y', '', '\uacf5\uacf5\ub370\uc774\ud130\ud3ec\ud138 \ubc1c\uae09\ud0a4'),
                    ('\uce21\uc815\uc18c\ucf54\ub4dc', 'site_id', 'VARCHAR', 20, 'N', 'N', '', ''),
                    ('\uc870\ud68c\uc2dc\uc791\uc77c', 'start_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
                    ('\uc870\ud68c\uc885\ub8cc\uc77c', 'end_date', 'CHAR', 8, 'N', 'N', '', 'YYYYMMDD'),
                ]
            },
            {
                'i_type': '\uad00\ub828\n\ud14c\uc774\ube14',
                'j_name': 'TB_COL_WATER_QUALITY',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uacb0\uacfc\ucf54\ub4dc', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                    ('\uce21\uc815\uc18c\uba85', 'site_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                    ('\uce21\uc815\uc77c\uc2dc', 'meas_date', 'DATETIME', 23, 'N', 'N', '', ''),
                    ('\uc218\uc18c\uc774\uc628\ub18d\ub3c4', 'ph', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: pH'),
                    ('\uc6a9\uc874\uc0b0\uc18c', 'do_value', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: mg/L'),
                ]
            },
            {
                'i_type': '\uad00\ub828\nAPI/URL',
                'j_name': 'https://apis.data.go.kr/B500001/rwis/waterQuality',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('BOD', 'bod', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: mg/L'),
                    ('\ud0c1\ub3c4', 'turbidity', 'NUMERIC', 8, 'Y', 'N', '', '\ub2e8\uc704: NTU'),
                    ('\uc794\ub958\uc5fc\uc18c', 'residual_cl', 'NUMERIC', 5, 'Y', 'N', '', '\ub2e8\uc704: mg/L'),
                ]
            },
        ]
    },
    {
        'area': '\ub370\uc774\ud130\ubc30\ud3ec',
        'if_id': 'IF-DH-004',
        'if_name': '\uad11\uc5ed\uc0c1\uc218\ub3c4 \uc720\ub7c9 \uc2e4\uc2dc\uac04 API',
        'req_sys': '\uc678\ubd80\uae30\uad00/\uc2dc\uc2a4\ud15c',
        'res_sys': '\ub370\uc774\ud130\ud5c8\ube0c\ud3ec\ud138',
        'in_out': '\uc678\ubd80',
        'cycle': '\uc2e4\uc2dc\uac04',
        'method': 'HTTP',
        'sections': [
            {
                'i_type': '\uad00\ub828\n\ud654\uba74',
                'j_name': 'Data Product \ubaa9\ub85d',
                'k_dir': '\uc1a1\uc2e0',
                'fields': [
                    ('API\uc778\uc99d\ud0a4', 'api_key', 'VARCHAR', 200, 'N', 'Y', '', ''),
                    ('\uc9c0\uc5ed\ucf54\ub4dc', 'region_code', 'VARCHAR', 10, 'Y', 'N', '', ''),
                    ('\uc2dc\uc124\ucf54\ub4dc', 'facility_id', 'VARCHAR', 20, 'Y', 'N', '', ''),
                    ('\uc870\ud68c\uc2dc\uc791\uc77c\uc2dc', 'from_date', 'DATETIME', 23, 'N', 'N', '', 'from'),
                    ('\uc870\ud68c\uc885\ub8cc\uc77c\uc2dc', 'to_date', 'DATETIME', 23, 'N', 'N', '', 'to'),
                ]
            },
            {
                'i_type': '\uad00\ub828\n\ud14c\uc774\ube14',
                'j_name': 'TB_DIST_FLOW_REALTIME',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uc751\ub2f5\ucf54\ub4dc', 'status_code', 'CHAR', 3, 'N', 'N', '', 'HTTP \uc0c1\ud0dc\ucf54\ub4dc'),
                    ('\ucd1d\uac74\uc218', 'total_count', 'NUMERIC', 10, 'N', 'N', '', ''),
                    ('\uc2dc\uc124\uba85', 'facility_name', 'VARCHAR', 100, 'N', 'N', '', ''),
                    ('\uce21\uc815\uc77c\uc2dc', 'meas_datetime', 'DATETIME', 23, 'N', 'Y', '', 'PK'),
                ]
            },
            {
                'i_type': '\uad00\ub828\nAPI/URL',
                'j_name': '/api/v2/water/flow/realtime',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uc720\ub7c9', 'flow_rate', 'NUMERIC', 10, 'N', 'N', '', '\ub2e8\uc704: m\u00b3/s'),
                    ('\uc218\uc555', 'pressure', 'NUMERIC', 8, 'Y', 'N', '', '\ub2e8\uc704: kPa'),
                    ('\uc218\uc704', 'water_level', 'NUMERIC', 8, 'Y', 'N', '', '\ub2e8\uc704: m'),
                ]
            },
        ]
    },
    {
        'area': '\uc2dc\uc2a4\ud15c\uad00\ub9ac',
        'if_id': 'IF-DH-005',
        'if_name': 'SSO \ud1b5\ud569\uc778\uc99d \uc0ac\uc6a9\uc790 \uc5f0\uacc4',
        'req_sys': '\ub370\uc774\ud130\ud5c8\ube0c\ud3ec\ud138',
        'res_sys': 'K-water SSO \uc778\uc99d\uc11c\ubc84',
        'in_out': '\ub0b4\ubd80',
        'cycle': '\uc2e4\uc2dc\uac04',
        'method': 'HTTP',
        'sections': [
            {
                'i_type': '\uad00\ub828\n\ud654\uba74',
                'j_name': '\ub85c\uadf8\uc778 \ud654\uba74',
                'k_dir': '\uc1a1\uc2e0',
                'fields': [
                    ('SAML\uc694\uccad', 'saml_request', 'TEXT', '', 'N', 'N', '', 'Base64 \uc778\ucf54\ub529'),
                    ('\ub9ac\ud134URL', 'relay_state', 'VARCHAR', 500, 'Y', 'N', '', '\uc778\uc99d \ud6c4 \ub9ac\ud134 \uc8fc\uc18c'),
                    ('\uc778\uc99d\ubc29\uc2dd', 'auth_method', 'CHAR', 3, 'N', 'N', 'SSO', 'SSO/PWD'),
                ]
            },
            {
                'i_type': '\uad00\ub828\n\ud14c\uc774\ube14',
                'j_name': 'TB_SYS_SESSION',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uc0ac\ubc88', 'emp_no', 'VARCHAR', 20, 'N', 'Y', '', 'PK'),
                    ('\uc0ac\uc6a9\uc790\uba85', 'user_name', 'VARCHAR', 50, 'N', 'N', '', ''),
                    ('\ubd80\uc11c\ucf54\ub4dc', 'dept_code', 'VARCHAR', 20, 'N', 'N', '', ''),
                    ('\uc5ed\ud560\ucf54\ub4dc', 'role_code', 'VARCHAR', 20, 'N', 'N', '', '\uad00\ub9ac\uc790/\uc77c\ubc18/\uc5d4\uc9c0\ub2c8\uc5b4'),
                ]
            },
            {
                'i_type': '\uad00\ub828\nAPI/URL',
                'j_name': 'https://sso.kwater.or.kr/auth/saml',
                'k_dir': '\uc218\uc2e0',
                'fields': [
                    ('\uc778\uc99d\ud1a0\ud070', 'auth_token', 'VARCHAR', 500, 'N', 'N', '', 'JWT \ud1a0\ud070'),
                    ('\ub9cc\ub8cc\uc77c\uc2dc', 'token_expire', 'DATETIME', 23, 'N', 'N', '', ''),
                    ('\uacb0\uacfc\ucf54\ub4dc', 'result_code', 'CHAR', 5, 'N', 'N', '', ''),
                ]
            },
        ]
    },
]

row = 4
for iface in interfaces:
    row = write_interface(ws, row, iface)

ws_cover = wb.worksheets[0]
ws_cover['C17'] = 'KWDP-DH-DS-17'
ws_cover['C19'] = '0.1'

wb.save(dst)
print(f'Saved: {dst}')
print(f'Total {row - 4} rows (row 4 ~ row {row - 1})')
