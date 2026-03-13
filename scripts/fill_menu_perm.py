import sys, io, copy
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Alignment, Font

src = r'C:/Users/yhw/OneDrive/바탕 화면/산출물/화면설계서/최종/KWDP-DH-DS-11-화면목록-v0.6.xlsx'
wb = openpyxl.load_workbook(src)
menu = wb.worksheets[2]

role_cols = {
    8:  '수공직원|일반사용자',
    9:  '수공직원|부서관리자',
    10: '수공직원|데이터스튜어드',
    11: '협력직원|데이터조회',
    12: '협력직원|외부개발자',
    13: '데이터엔지니어|파이프라인관리자',
    14: '데이터엔지니어|ETL운영자',
    15: '데이터엔지니어|DBA',
    16: '관리자|시스템관리자',
    17: '관리자|보안관리자',
    18: '관리자|슈퍼관리자',
}

RBAC = {
    '수공직원|일반사용자': [
        'home', 'process', 'quality', 'measurement', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'dist-product', 'dist-approval', 'dist-stats',
        'comm-notice', 'comm-internal', 'comm-archive'
    ],
    '수공직원|부서관리자': [
        'home', 'process', 'quality', 'measurement', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-arch', 'col-monitor',
        'dist-product', 'dist-deidentify', 'dist-approval', 'dist-api', 'dist-external', 'dist-stats',
        'meta-glossary', 'meta-tag', 'meta-model', 'meta-dq',
        'meta-std-dashboard', 'meta-std-words', 'meta-std-domains', 'meta-std-terms', 'meta-std-codes', 'meta-std-import',
        'comm-notice', 'comm-internal', 'comm-archive'
    ],
    '수공직원|데이터스튜어드': [
        'home', 'process', 'quality', 'measurement', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-dbt',
        'dist-product', 'dist-deidentify', 'dist-approval', 'dist-api', 'dist-external', 'dist-stats', 'dist-chart-content',
        'meta-glossary', 'meta-tag', 'meta-model', 'meta-dq', 'meta-ontology',
        'meta-std-dashboard', 'meta-std-words', 'meta-std-domains', 'meta-std-terms', 'meta-std-codes', 'meta-std-import',
        'sys-security',
        'comm-notice', 'comm-internal', 'comm-archive'
    ],
    '협력직원|데이터조회': [
        'home', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'dist-approval',
        'comm-notice', 'comm-external', 'comm-archive'
    ],
    '협력직원|외부개발자': [
        'home', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'dist-approval', 'dist-api',
        'comm-notice', 'comm-external', 'comm-archive'
    ],
    '데이터엔지니어|파이프라인관리자': [
        'home', 'process', 'quality', 'measurement', 'asset-db', 'digital-twin', 'ai', 'llmops', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'meta-dq',
        'dist-product', 'dist-approval', 'dist-stats',
        'sys-interface', 'sys-k8s',
        'comm-notice', 'comm-internal', 'comm-archive'
    ],
    '데이터엔지니어|ETL운영자': [
        'home', 'process', 'quality', 'measurement', 'asset-db', 'digital-twin', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'meta-dq',
        'dist-product', 'dist-approval', 'dist-stats',
        'sys-interface',
        'comm-notice', 'comm-internal', 'comm-archive'
    ],
    '데이터엔지니어|DBA': [
        'home', 'process', 'quality', 'measurement', 'asset-db', 'digital-twin', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'meta-glossary', 'meta-tag', 'meta-model', 'meta-dq', 'meta-ontology',
        'meta-std-dashboard', 'meta-std-words', 'meta-std-domains', 'meta-std-terms', 'meta-std-codes', 'meta-std-import',
        'dist-product', 'dist-approval', 'dist-stats',
        'comm-notice', 'comm-internal', 'comm-archive'
    ],
    '관리자|시스템관리자': [
        'home', 'process', 'quality', 'measurement', 'asset-db', 'digital-twin', 'ai', 'llmops', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'dist-product', 'dist-approval', 'dist-stats', 'dist-chart-content',
        'sys-user', 'sys-role', 'sys-security', 'sys-interface', 'sys-audit', 'sys-widget-template', 'sys-perm', 'sys-engine', 'sys-k8s', 'sys-erp-sync', 'sys-ext-register',
        'comm-notice', 'comm-internal', 'comm-external', 'comm-archive'
    ],
    '관리자|보안관리자': [
        'home', 'process', 'quality', 'measurement', 'asset-db', 'digital-twin', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-arch', 'col-log',
        'dist-deidentify', 'dist-approval', 'dist-stats', 'dist-chart-content',
        'sys-user', 'sys-role', 'sys-security', 'sys-audit', 'sys-widget-template', 'sys-perm', 'sys-engine', 'sys-k8s', 'sys-erp-sync', 'sys-ext-register',
        'comm-notice', 'comm-archive'
    ],
    '관리자|슈퍼관리자': 'ALL',
}

SCREEN_PARENT_MAP = {
    'col-system': 'col-pipeline', 'col-system-detail': 'col-pipeline', 'sys-interface-detail': 'sys-interface',
    'dist-register': 'dist-product', 'dist-detail': 'dist-product',
    'dist-deidentify-detail': 'dist-deidentify', 'dist-api-detail': 'dist-api',
    'meta-model-detail': 'meta-model', 'meta-dq-detail': 'meta-dq',
    'gre-office': 'measurement', 'gre-site': 'measurement',
    'cat-ontology': 'meta-ontology', 'col-dbt-detail': 'col-dbt',
    'cat-data-view': 'cat-lineage', 'cat-quality-report': 'cat-lineage',
    'new-request': 'process',
    'sys-widget-template-register': 'sys-widget-template', 'sys-widget-template-detail': 'sys-widget-template',
    'dist-chart-content-register': 'dist-chart-content', 'dist-chart-content-detail': 'dist-chart-content',
    'cat-detail': 'cat-search', 'col-register': 'col-pipeline',
    'sys-user-detail': 'sys-user',
    'sys-ext-register-detail': 'sys-ext-register',
}

MENU_SCREEN_MAP = {
    'DH0101': ['home'], 'DH0102': ['sitemap'],
    'DH010301': ['home'], 'DH010302': ['home'], 'DH010303': ['home'],
    'DH0201': ['home'], 'DH0202': ['process'], 'DH020201': ['new-request'],
    'DH0203': ['quality'], 'DH0204': ['ai'], 'DH0205': ['gallery'],
    'DH0206': ['widget-settings'],
    'DH0207': ['dist-chart-content'], 'DH020701': ['dist-chart-content-register'],
    'DH0208': ['sys-widget-template'],
    'DH020801': ['sys-widget-template-register'], 'DH020802': ['sys-widget-template-detail'],
    'DH0301': ['measurement'], 'DH030101': ['gre-office'], 'DH030102': ['gre-site'],
    'DH0302': ['asset-db'],
    'DH0401': ['cat-search'], 'DH040101': ['cat-detail'],
    'DH0402': ['cat-graph'], 'DH0403': ['cat-lineage'],
    'DH040301': ['cat-data-view'], 'DH040302': ['cat-quality-report'],
    'DH0404': ['cat-bookmark'], 'DH0405': ['cat-request'],
    'DH0501': ['col-pipeline'], 'DH050101': ['col-register'],
    'DH050102': ['col-system'], 'DH050103': ['col-system-detail'],
    'DH0502': ['col-monitor'], 'DH0503': ['col-cdc'],
    'DH0504': ['col-kafka'], 'DH050401': ['col-kafka'], 'DH050402': ['col-kafka'],
    'DH0505': ['col-external'],
    'DH0506': ['col-arch'], 'DH050601': ['col-arch'], 'DH050602': ['col-arch'],
    'DH0507': ['col-log'],
    'DH0508': ['col-dbt'], 'DH050801': ['col-dbt-detail'],
    'DH0601': ['dist-product'], 'DH060101': ['dist-register'],
    'DH060102': ['dist-detail'], 'DH060103': ['dist-product'],
    'DH060104': ['dist-api'], 'DH060105': ['dist-api'],
    'DH0602': ['dist-deidentify'],
    'DH060201': ['dist-deidentify-detail'], 'DH060202': ['dist-deidentify'], 'DH060203': ['dist-deidentify'],
    'DH0603': ['dist-approval'],
    'DH0604': ['dist-api'], 'DH060401': ['dist-api-detail'],
    'DH0605': ['dist-external'], 'DH0606': ['dist-stats'],
    'DH0701': ['meta-glossary'], 'DH070101': ['meta-glossary'],
    'DH070102': ['meta-glossary'], 'DH070103': ['meta-glossary'], 'DH070104': ['meta-glossary'],
    'DH0702': ['meta-tag'], 'DH070201': ['meta-tag'],
    'DH0703': ['meta-model'], 'DH070301': ['meta-model-detail'],
    'DH070302': ['meta-model'], 'DH070303': ['meta-model'],
    'DH070304': ['meta-model'], 'DH070305': ['meta-model'],
    'DH0704': ['meta-dq'], 'DH070401': ['meta-dq-detail'],
    'DH070402': ['meta-dq'], 'DH070403': ['meta-dq'],
    'DH0705': ['meta-ontology'],
    'DH070501': ['meta-ontology'], 'DH070502': ['meta-ontology'],
    'DH070503': ['meta-ontology'], 'DH070504': ['meta-ontology'],
    'DH070505': ['meta-ontology'], 'DH070506': ['meta-ontology'], 'DH070507': ['meta-ontology'],
    'DH0801': ['comm-notice'], 'DH0802': ['comm-internal'],
    'DH0803': ['comm-external'], 'DH0804': ['comm-archive'],
    'DH0901': ['sys-user'], 'DH0902': ['sys-role'],
    'DH090201': ['sys-role'], 'DH090202': ['sys-role'],
    'DH090203': ['sys-role'], 'DH090204': ['sys-role'],
    'DH090205': ['sys-audit'],
    'DH0903': ['sys-security'], 'DH0904': ['sys-interface'],
    'DH0905': ['sys-audit'], 'DH0906': ['sys-perm'], 'DH0907': ['sys-engine'],
}

COMMON_MENUS = {'DH0101', 'DH0102', 'DH010301', 'DH010302', 'DH010303'}

def has_access(role_key, screen_ids):
    rbac = RBAC[role_key]
    if rbac == 'ALL':
        return True
    for sid in screen_ids:
        if sid in rbac:
            return True
        parent = SCREEN_PARENT_MAP.get(sid)
        if parent and parent in rbac:
            return True
    return False

filled_count = 0
for row in range(5, menu.max_row + 1):
    menu_id = None
    for col in [6, 4, 2]:
        val = menu.cell(row=row, column=col).value
        if val and str(val).startswith('DH'):
            menu_id = str(val)
            break
    if not menu_id:
        continue

    if menu_id in COMMON_MENUS:
        for col_idx in role_cols:
            cell = menu.cell(row=row, column=col_idx)
            cell.value = '\u25cf'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            filled_count += 1
    elif menu_id in MENU_SCREEN_MAP:
        screen_ids = MENU_SCREEN_MAP[menu_id]
        for col_idx, role_key in role_cols.items():
            if has_access(role_key, screen_ids):
                cell = menu.cell(row=row, column=col_idx)
                cell.value = '\u25cf'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                filled_count += 1

wb.save(src)
print(f'완료: {filled_count}개 권한 셀에 기호 입력됨')
