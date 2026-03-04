import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

# Load template
src = 'D:/00_수공프로젝트/20260224_설계단계_산출물/산출물 템플릿_데이터허브(손이사님 제공)/4. 데이터허브/2. 설계/(템플릿)KWDP-DH-DS-11-화면목록-v0.x.xlsx'
wb = openpyxl.load_workbook(src)

ws = wb[wb.sheetnames[2]]  # 메뉴구조도

# ===== RBAC DATA FROM PROJECT (main.js RBAC_MATRIX) =====
RBAC_MATRIX = {
    '수공직원|일반사용자': ['home', 'process', 'quality', 'measurement', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'dist-product', 'dist-approval', 'dist-stats',
        'comm-notice', 'comm-internal', 'comm-archive'],
    '수공직원|부서관리자': ['home', 'process', 'quality', 'measurement', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-arch', 'col-monitor',
        'dist-product', 'dist-deidentify', 'dist-approval', 'dist-api', 'dist-external', 'dist-stats',
        'meta-glossary', 'meta-tag', 'meta-model', 'meta-dq',
        'comm-notice', 'comm-internal', 'comm-archive'],
    '수공직원|데이터스튜어드': ['home', 'process', 'quality', 'measurement', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-dbt',
        'dist-product', 'dist-deidentify', 'dist-approval', 'dist-api', 'dist-external', 'dist-stats', 'dist-chart-content',
        'meta-glossary', 'meta-tag', 'meta-model', 'meta-dq', 'meta-ontology',
        'sys-security',
        'comm-notice', 'comm-internal', 'comm-archive'],
    '협력직원|데이터조회': ['home', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'dist-approval',
        'comm-notice', 'comm-external', 'comm-archive'],
    '협력직원|외부개발자': ['home', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'dist-approval', 'dist-api',
        'comm-notice', 'comm-external', 'comm-archive'],
    '데이터엔지니어|파이프라인관리자': ['home', 'process', 'quality', 'measurement', 'asset-db', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'meta-dq',
        'dist-product', 'dist-approval', 'dist-stats',
        'sys-interface',
        'comm-notice', 'comm-internal', 'comm-archive'],
    '데이터엔지니어|ETL운영자': ['home', 'process', 'quality', 'measurement', 'asset-db', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'meta-dq',
        'dist-product', 'dist-approval', 'dist-stats',
        'sys-interface',
        'comm-notice', 'comm-internal', 'comm-archive'],
    '데이터엔지니어|DBA': ['home', 'process', 'quality', 'measurement', 'asset-db', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'meta-glossary', 'meta-tag', 'meta-model', 'meta-dq', 'meta-ontology',
        'dist-product', 'dist-approval', 'dist-stats',
        'comm-notice', 'comm-internal', 'comm-archive'],
    '관리자|시스템관리자': ['home', 'process', 'quality', 'measurement', 'asset-db', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-pipeline', 'col-register', 'col-cdc', 'col-kafka', 'col-external', 'col-arch', 'col-monitor', 'col-log', 'col-dbt',
        'dist-product', 'dist-approval', 'dist-stats', 'dist-chart-content',
        'sys-user', 'sys-role', 'sys-security', 'sys-interface', 'sys-audit', 'sys-widget-template', 'sys-perm', 'sys-engine',
        'comm-notice', 'comm-internal', 'comm-external', 'comm-archive'],
    '관리자|보안관리자': ['home', 'process', 'quality', 'measurement', 'asset-db', 'ai', 'gallery', 'widget-settings', 'sitemap',
        'cat-search', 'cat-detail', 'cat-graph', 'cat-lineage', 'cat-bookmark', 'cat-request',
        'col-arch', 'col-log',
        'dist-deidentify', 'dist-approval', 'dist-stats', 'dist-chart-content',
        'sys-user', 'sys-role', 'sys-security', 'sys-audit', 'sys-widget-template', 'sys-perm',
        'comm-notice', 'comm-archive'],
    '관리자|슈퍼관리자': 'ALL'
}

# ===== 전체 11개 역할 (열 순서) =====
# 그룹 > 세부역할 순서
ALL_ROLES = [
    # 수공직원 (3)
    ('수공직원', '일반사용자', '수공직원|일반사용자'),
    ('수공직원', '부서관리자', '수공직원|부서관리자'),
    ('수공직원', '데이터\n스튜어드', '수공직원|데이터스튜어드'),
    # 협력직원 (2)
    ('협력직원', '데이터조회', '협력직원|데이터조회'),
    ('협력직원', '외부개발자', '협력직원|외부개발자'),
    # 데이터엔지니어 (3)
    ('데이터엔지니어', '파이프라인\n관리자', '데이터엔지니어|파이프라인관리자'),
    ('데이터엔지니어', 'ETL\n운영자', '데이터엔지니어|ETL운영자'),
    ('데이터엔지니어', 'DBA', '데이터엔지니어|DBA'),
    # 관리자 (3)
    ('관리자', '시스템\n관리자', '관리자|시스템관리자'),
    ('관리자', '보안\n관리자', '관리자|보안관리자'),
    ('관리자', '슈퍼\n관리자', '관리자|슈퍼관리자'),
]

# 그룹 헤더 정보: (그룹명, 시작인덱스, 개수)
ROLE_GROUP_HEADERS = [
    ('수공직원', 0, 3),
    ('협력직원', 3, 2),
    ('데이터엔지니어', 5, 3),
    ('관리자', 8, 3),
]

PERM_COL_START = 8  # H열부터 권한 시작
REMARK_COL = PERM_COL_START + len(ALL_ROLES)  # 비고 열 = 19 (S열)
TOTAL_COLS = REMARK_COL  # 19

# SCREEN_PARENT_MAP: child screen inherits parent access
SCREEN_PARENT_MAP = {
    'col-system': 'col-pipeline', 'col-system-detail': 'col-pipeline',
    'sys-interface-detail': 'sys-interface',
    'dist-register': 'dist-product', 'dist-detail': 'dist-product',
    'dist-deidentify-detail': 'dist-deidentify', 'dist-api-detail': 'dist-api',
    'meta-model-detail': 'meta-model', 'meta-dq-detail': 'meta-dq',
    'gre-office': 'measurement', 'gre-site': 'measurement',
    'cat-ontology': 'meta-ontology', 'col-dbt-detail': 'col-dbt',
    'cat-data-view': 'cat-lineage', 'cat-quality-report': 'cat-lineage',
    'new-request': 'process',
    'sys-widget-template-register': 'sys-widget-template', 'sys-widget-template-detail': 'sys-widget-template',
    'dist-chart-content-register': 'dist-chart-content', 'dist-chart-content-detail': 'dist-chart-content',
    'cat-detail': 'cat-search', 'col-register': 'col-pipeline'
}

def has_role_access(screen_id, role_key):
    """개별 역할의 화면 접근 여부 확인 (부모 화면 상속 포함)"""
    screens = RBAC_MATRIX[role_key]
    if screens == 'ALL' or screen_id in screens:
        return True
    parent = SCREEN_PARENT_MAP.get(screen_id)
    if parent:
        if screens == 'ALL' or parent in screens:
            return True
    return False

# ===== SCREEN HIERARCHY =====
SCREENS = [
    # (Level1_ID, Level1_Name, Level2_ID, Level2_Name, Level3_ID, Level3_Name, screen_id)
    ('DH01', '개인 대시보드', 'DH0101', '개인대시보드', '', '', 'home'),
    ('DH01', '개인 대시보드', 'DH0102', '데이터 워크플로우', '', '', 'process'),
    ('DH01', '개인 대시보드', 'DH0102', '데이터 워크플로우', 'DH010201', '신규요청 등록', 'new-request'),
    ('DH01', '개인 대시보드', 'DH0103', '데이터 품질현황 종합', '', '', 'quality'),
    ('DH01', '개인 대시보드', 'DH0104', 'AI 검색 및 질의응답', '', '', 'ai'),
    ('DH01', '개인 대시보드', 'DH0105', '시각화·차트 갤러리', '', '', 'gallery'),
    ('DH01', '개인 대시보드', 'DH0106', '위젯 설정', '', '', 'widget-settings'),
    ('DH01', '개인 대시보드', 'DH0107', '차트 콘텐츠 관리', '', '', 'dist-chart-content'),
    ('DH01', '개인 대시보드', 'DH0107', '차트 콘텐츠 관리', 'DH010701', '차트 콘텐츠 등록', 'dist-chart-content-register'),
    ('DH01', '개인 대시보드', 'DH0107', '차트 콘텐츠 관리', 'DH010702', '차트 콘텐츠 상세', 'dist-chart-content-detail'),
    ('DH01', '개인 대시보드', 'DH0108', '위젯 템플릿 관리', '', '', 'sys-widget-template'),
    ('DH01', '개인 대시보드', 'DH0108', '위젯 템플릿 관리', 'DH010801', '위젯 템플릿 등록', 'sys-widget-template-register'),
    ('DH01', '개인 대시보드', 'DH0108', '위젯 템플릿 관리', 'DH010802', '위젯 템플릿 상세', 'sys-widget-template-detail'),

    ('DH02', '모니터링', 'DH0201', '실시간 계측DB 모니터링', '', '', 'measurement'),
    ('DH02', '모니터링', 'DH0201', '실시간 계측DB 모니터링', 'DH020101', '본사 그린에너지 상세', 'gre-office'),
    ('DH02', '모니터링', 'DH0201', '실시간 계측DB 모니터링', 'DH020102', '현장 그린에너지 상세', 'gre-site'),
    ('DH02', '모니터링', 'DH0202', '자산DB 모니터링', '', '', 'asset-db'),

    ('DH03', '카탈로그·탐색', 'DH0301', '통합데이터 검색', '', '', 'cat-search'),
    ('DH03', '카탈로그·탐색', 'DH0301', '통합데이터 검색', 'DH030101', '데이터셋 상세', 'cat-detail'),
    ('DH03', '카탈로그·탐색', 'DH0302', '지식그래프·온톨로지', '', '', 'cat-graph'),
    ('DH03', '카탈로그·탐색', 'DH0303', '데이터 리니지(계보)', '', '', 'cat-lineage'),
    ('DH03', '카탈로그·탐색', 'DH0303', '데이터 리니지(계보)', 'DH030301', '데이터 뷰', 'cat-data-view'),
    ('DH03', '카탈로그·탐색', 'DH0303', '데이터 리니지(계보)', 'DH030302', '품질 리포트', 'cat-quality-report'),
    ('DH03', '카탈로그·탐색', 'DH0304', '내 보관함', '', '', 'cat-bookmark'),
    ('DH03', '카탈로그·탐색', 'DH0305', '데이터 요청·피드백', '', '', 'cat-request'),

    ('DH04', '데이터 수집·통합', 'DH0401', '파이프라인 관리', '', '', 'col-pipeline'),
    ('DH04', '데이터 수집·통합', 'DH0401', '파이프라인 관리', 'DH040101', '파이프라인 등록', 'col-register'),
    ('DH04', '데이터 수집·통합', 'DH0401', '파이프라인 관리', 'DH040102', '소스시스템 관리', 'col-system'),
    ('DH04', '데이터 수집·통합', 'DH0401', '파이프라인 관리', 'DH040103', '소스시스템 상세', 'col-system-detail'),
    ('DH04', '데이터 수집·통합', 'DH0402', '수집현황 모니터링', '', '', 'col-monitor'),
    ('DH04', '데이터 수집·통합', 'DH0403', 'CDC 연계현황', '', '', 'col-cdc'),
    ('DH04', '데이터 수집·통합', 'DH0404', 'Kafka 스트리밍', '', '', 'col-kafka'),
    ('DH04', '데이터 수집·통합', 'DH0405', '외부연계', '', '', 'col-external'),
    ('DH04', '데이터 수집·통합', 'DH0406', '연계아키텍처 구조도', '', '', 'col-arch'),
    ('DH04', '데이터 수집·통합', 'DH0407', '실시간 로그·알림', '', '', 'col-log'),
    ('DH04', '데이터 수집·통합', 'DH0408', '정제·변환 관리(dbt)', '', '', 'col-dbt'),
    ('DH04', '데이터 수집·통합', 'DH0408', '정제·변환 관리(dbt)', 'DH040801', 'dbt 모델 상세', 'col-dbt-detail'),

    ('DH05', '데이터 유통·활용', 'DH0501', 'Data Product 구성·관리', '', '', 'dist-product'),
    ('DH05', '데이터 유통·활용', 'DH0501', 'Data Product 구성·관리', 'DH050101', 'Data Product 등록', 'dist-register'),
    ('DH05', '데이터 유통·활용', 'DH0501', 'Data Product 구성·관리', 'DH050102', 'Data Product 상세', 'dist-detail'),
    ('DH05', '데이터 유통·활용', 'DH0502', '비식별화·가명처리', '', '', 'dist-deidentify'),
    ('DH05', '데이터 유통·활용', 'DH0502', '비식별화·가명처리', 'DH050201', '비식별화 정책 상세', 'dist-deidentify-detail'),
    ('DH05', '데이터 유통·활용', 'DH0503', '데이터 신청·승인', '', '', 'dist-approval'),
    ('DH05', '데이터 유통·활용', 'DH0504', '유통 API 관리', '', '', 'dist-api'),
    ('DH05', '데이터 유통·활용', 'DH0504', '유통 API 관리', 'DH050401', 'API 상세', 'dist-api-detail'),
    ('DH05', '데이터 유통·활용', 'DH0505', '외부시스템 데이터 제공', '', '', 'dist-external'),
    ('DH05', '데이터 유통·활용', 'DH0506', '데이터 유통통계', '', '', 'dist-stats'),

    ('DH06', '메타데이터 관리', 'DH0601', '표준용어·코드관리', '', '', 'meta-glossary'),
    ('DH06', '메타데이터 관리', 'DH0602', '분류체계·태그관리', '', '', 'meta-tag'),
    ('DH06', '메타데이터 관리', 'DH0603', '데이터모델관리', '', '', 'meta-model'),
    ('DH06', '메타데이터 관리', 'DH0603', '데이터모델관리', 'DH060301', '데이터모델 상세', 'meta-model-detail'),
    ('DH06', '메타데이터 관리', 'DH0604', 'DQ 검증·프로파일링', '', '', 'meta-dq'),
    ('DH06', '메타데이터 관리', 'DH0604', 'DQ 검증·프로파일링', 'DH060401', 'DQ 규칙 상세', 'meta-dq-detail'),
    ('DH06', '메타데이터 관리', 'DH0605', '온톨로지 관리', '', '', 'meta-ontology'),

    ('DH07', '커뮤니티·소통', 'DH0701', '공지사항', '', '', 'comm-notice'),
    ('DH07', '커뮤니티·소통', 'DH0702', '내부게시판', '', '', 'comm-internal'),
    ('DH07', '커뮤니티·소통', 'DH0703', '외부게시판', '', '', 'comm-external'),
    ('DH07', '커뮤니티·소통', 'DH0704', '자료실', '', '', 'comm-archive'),

    ('DH08', '시스템관리', 'DH0801', '조직 및 사용자관리', '', '', 'sys-user'),
    ('DH08', '시스템관리', 'DH0802', '권한 및 역할관리', '', '', 'sys-role'),
    ('DH08', '시스템관리', 'DH0803', '데이터등급·보안정책', '', '', 'sys-security'),
    ('DH08', '시스템관리', 'DH0804', '연계인터페이스 모니터링', '', '', 'sys-interface'),
    ('DH08', '시스템관리', 'DH0804', '연계인터페이스 모니터링', 'DH080401', '인터페이스 상세', 'sys-interface-detail'),
    ('DH08', '시스템관리', 'DH0805', '접속통계·감사로그', '', '', 'sys-audit'),
    ('DH08', '시스템관리', 'DH0806', '화면별 권한설정', '', '', 'sys-perm'),
    ('DH08', '시스템관리', 'DH0807', '데이터허브 엔진관리', '', '', 'sys-engine'),

    ('DH09', '공통', 'DH0901', '사이트맵', '', '', 'sitemap'),
    ('DH09', '공통', 'DH0902', '온톨로지 클래스탐색', '', '', 'cat-ontology'),
]

# ===== STYLES =====
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill_blue = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
# 그룹별 색상
group_fills = {
    '수공직원':       PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),  # 연한 파랑
    '협력직원':       PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # 연한 초록
    '데이터엔지니어': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),  # 연한 주황
    '관리자':         PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),  # 연한 보라
}
group_header_fills = {
    '수공직원':       PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'),
    '협력직원':       PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),
    '데이터엔지니어': PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid'),
    '관리자':         PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),
}
sub_header_font = Font(name='맑은 고딕', size=8, bold=True)
data_font = Font(name='맑은 고딕', size=9)
data_font_small = Font(name='맑은 고딕', size=8)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
perm_font = Font(name='맑은 고딕', size=10, bold=True, color='2F5496')

# ===== CLEAR ENTIRE SHEET =====
# Unmerge first (before clearing cell values)
for merged in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merged))

# Then clear all cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=30):
    for cell in row:
        cell.value = None
        cell.border = Border()
        cell.fill = PatternFill()
        cell.font = Font()
        cell.alignment = Alignment()

# ===== BUILD HEADER =====
# Row 1: Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
title_cell = ws.cell(row=1, column=1, value='데이터허브 포털 - 화면목록')
title_cell.font = Font(name='맑은 고딕', size=14, bold=True, color='1F3864')
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Row 2: Column group headers
# A2: 번호
ws.cell(row=2, column=1, value='번호')
ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
# B2~C2: Level 1
ws.cell(row=2, column=2, value='Level 1')
ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
# D2~E2: Level 2
ws.cell(row=2, column=4, value='Level 2')
ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
# F2~G2: Level 3
ws.cell(row=2, column=6, value='Level 3')
ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)

# Row 2: Role group headers (merged)
for group_name, start_idx, count in ROLE_GROUP_HEADERS:
    col_start = PERM_COL_START + start_idx
    col_end = col_start + count - 1
    ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
    cell = ws.cell(row=2, column=col_start, value=group_name)
    cell.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
    cell.alignment = center_align
    cell.fill = group_header_fills[group_name]
    # Apply border+fill to all cells in merged range
    for c in range(col_start, col_end + 1):
        mc = ws.cell(row=2, column=c)
        mc.border = thin_border
        mc.fill = group_header_fills[group_name]

# Remark column header
ws.cell(row=2, column=REMARK_COL, value='비고')
ws.merge_cells(start_row=2, start_column=REMARK_COL, end_row=3, end_column=REMARK_COL)

# Style Row 2 (A~G)
for col_idx in range(1, 8):
    cell = ws.cell(row=2, column=col_idx)
    cell.font = header_font_white
    cell.alignment = center_align
    cell.fill = header_fill_blue
    cell.border = thin_border

# Row 2 remark
remark_cell = ws.cell(row=2, column=REMARK_COL)
remark_cell.font = header_font_white
remark_cell.alignment = center_align
remark_cell.fill = header_fill_blue
remark_cell.border = thin_border

# Row 3: Sub-headers
sub_labels = ['', '메뉴ID', '메뉴명', '메뉴ID', '메뉴명', '메뉴ID', '메뉴명']
for col_idx, label in enumerate(sub_labels):
    if col_idx == 0:
        continue  # 번호는 Row2에서 merge됨
    cell = ws.cell(row=3, column=col_idx + 1, value=label)
    cell.font = sub_header_font
    cell.alignment = center_align
    cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    cell.border = thin_border

# Row 3: Individual role names
for i, (group_name, role_label, role_key) in enumerate(ALL_ROLES):
    col_idx = PERM_COL_START + i
    cell = ws.cell(row=3, column=col_idx, value=role_label)
    cell.font = Font(name='맑은 고딕', size=7, bold=True)
    cell.alignment = center_align
    cell.fill = group_fills[group_name]
    cell.border = thin_border

# Row 3: 번호(merged), 비고(merged) - already merged, just add border
ws.cell(row=3, column=1).border = thin_border
ws.cell(row=3, column=1).fill = header_fill_blue
ws.cell(row=3, column=REMARK_COL).border = thin_border
ws.cell(row=3, column=REMARK_COL).fill = header_fill_blue

# ===== WRITE DATA ROWS =====
DATA_START_ROW = 4
row_num = DATA_START_ROW
prev_l1_id = ''
prev_l2_id = ''

# 그룹 구분 색상 (Level 1 그룹별 교대)
l1_fills = [
    PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),  # 흰색
    PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),  # 연한 회색
]
l1_idx = -1
last_l1 = ''

for l1_id, l1_name, l2_id, l2_name, l3_id, l3_name, screen_id in SCREENS:
    # Level 1 그룹 교대 색상
    if l1_id != last_l1:
        l1_idx += 1
        last_l1 = l1_id
    row_fill = l1_fills[l1_idx % 2]

    # Number
    ws.cell(row=row_num, column=1, value=row_num - DATA_START_ROW + 1)

    # Level 1
    if l1_id != prev_l1_id:
        ws.cell(row=row_num, column=2, value=l1_id)
        ws.cell(row=row_num, column=3, value=l1_name)
    prev_l1_id = l1_id

    # Level 2
    if l2_id != prev_l2_id:
        ws.cell(row=row_num, column=4, value=l2_id)
        ws.cell(row=row_num, column=5, value=l2_name)
    prev_l2_id = l2_id

    # Level 3
    if l3_id:
        ws.cell(row=row_num, column=6, value=l3_id)
        ws.cell(row=row_num, column=7, value=l3_name)

    # Permissions - 11개 역할 개별 체크
    for i, (group_name, role_label, role_key) in enumerate(ALL_ROLES):
        col_idx = PERM_COL_START + i
        if has_role_access(screen_id, role_key):
            cell = ws.cell(row=row_num, column=col_idx, value='\u25cf')
            cell.font = perm_font
        else:
            cell = ws.cell(row=row_num, column=col_idx, value='\u2013')
            cell.font = Font(name='맑은 고딕', size=9, color='C0C0C0')
        cell.alignment = center_align

    # Remarks
    if l3_id:
        remark = f'상세/등록 (screen: {screen_id})'
    else:
        remark = f'screen: {screen_id}'
    ws.cell(row=row_num, column=REMARK_COL, value=remark)

    # Apply styling to all cells
    for col_idx in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = thin_border
        # 배경색
        if PERM_COL_START <= col_idx < PERM_COL_START + len(ALL_ROLES):
            pass  # 권한 칸은 배경 유지
        else:
            cell.fill = row_fill
        # 폰트 (이미 설정된 권한 셀 제외)
        if col_idx < PERM_COL_START or col_idx >= PERM_COL_START + len(ALL_ROLES):
            if not cell.font or cell.font.size is None:
                cell.font = data_font
        # 정렬
        if col_idx in [1, 2, 4, 6] or (PERM_COL_START <= col_idx < PERM_COL_START + len(ALL_ROLES)):
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    row_num += 1

# ===== COLUMN WIDTHS =====
col_widths = {
    1: 5,    # 번호
    2: 7,    # L1 ID
    3: 16,   # L1 Name
    4: 9,    # L2 ID
    5: 22,   # L2 Name
    6: 10,   # L3 ID
    7: 20,   # L3 Name
    REMARK_COL: 24,  # 비고
}
# 권한 열 너비
for i in range(len(ALL_ROLES)):
    col_widths[PERM_COL_START + i] = 8

for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ===== SUMMARY ROW =====
summary_row = row_num + 1
ws.cell(row=summary_row, column=1, value='합계')
ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=7)
summary_cell = ws.cell(row=summary_row, column=1)
summary_cell.font = Font(name='맑은 고딕', size=9, bold=True)
summary_cell.alignment = Alignment(horizontal='right', vertical='center')
summary_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

for i, (group_name, role_label, role_key) in enumerate(ALL_ROLES):
    col_idx = PERM_COL_START + i
    count = sum(1 for s in SCREENS if has_role_access(s[6], role_key))
    cell = ws.cell(row=summary_row, column=col_idx, value=count)
    cell.font = Font(name='맑은 고딕', size=9, bold=True)
    cell.alignment = center_align
    cell.fill = group_fills[group_name]
    cell.border = thin_border

ws.cell(row=summary_row, column=REMARK_COL, value=f'총 {len(SCREENS)}개 화면')
ws.cell(row=summary_row, column=REMARK_COL).font = Font(name='맑은 고딕', size=9, bold=True)
ws.cell(row=summary_row, column=REMARK_COL).alignment = left_align
ws.cell(row=summary_row, column=REMARK_COL).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
ws.cell(row=summary_row, column=REMARK_COL).border = thin_border

# Border for merged summary cells
for c in range(1, 8):
    ws.cell(row=summary_row, column=c).border = thin_border
    ws.cell(row=summary_row, column=c).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

# ===== PRINT SETTINGS =====
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# Freeze panes
ws.freeze_panes = 'A4'

# ===== UPDATE COVER PAGE =====
ws_cover = wb[wb.sheetnames[0]]
ws_cover['C17'] = 'KWDP-DH-DS-11'
ws_cover['C19'] = '0.9'
ws_cover['C21'] = '데이터허브 포털'

# ===== SAVE =====
out_path = 'D:/00_수공프로젝트/20260224_설계단계_산출물/산출물 템플릿_데이터허브(손이사님 제공)/4. 데이터허브/2. 설계/KWDP-DH-DS-11-화면목록-v1.0.xlsx'
wb.save(out_path)
print(f'파일 저장 완료: {out_path}')
print(f'총 {row_num - DATA_START_ROW}개 화면 항목 작성')

# Print summary
print('\n=== 화면 목록 요약 ===')
groups = {}
for l1_id, l1_name, l2_id, l2_name, l3_id, l3_name, screen_id in SCREENS:
    key = f'{l1_id} {l1_name}'
    if key not in groups:
        groups[key] = 0
    groups[key] += 1
for g, cnt in groups.items():
    print(f'  {g}: {cnt}개 화면')

print('\n=== 역할별 접근 가능 화면 수 (전체 11개 역할) ===')
for group_name, role_label, role_key in ALL_ROLES:
    count = sum(1 for s in SCREENS if has_role_access(s[6], role_key))
    print(f'  {group_name} > {role_label.replace(chr(10)," ")}: {count}/{len(SCREENS)}개')
