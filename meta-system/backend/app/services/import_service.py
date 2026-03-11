"""
엑셀 임포트/엑스포트 비즈니스 로직
"""
import io
from typing import Optional
from openpyxl import Workbook, load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import StreamingResponse

from app.models.word import StdWord
from app.models.domain import StdDomainGroup, StdDomain
from app.models.term import StdTerm
from app.models.code import StdCodeGroup, StdCode

BATCH_SIZE = 5000


def _parse_bool(value) -> bool:
    """Y/N/빈값을 boolean으로 변환"""
    if value is None or value == "":
        return False
    return str(value).strip().upper() == "Y"


def _parse_int(value) -> Optional[int]:
    """float/빈값을 int로 변환"""
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _clean_str(value) -> Optional[str]:
    """빈문자열을 None으로 변환"""
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip()


async def import_word_dict(db: AsyncSession, file_bytes: bytes) -> dict:
    """단어사전 엑셀 임포트"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    success_count = 0
    error_count = 0
    errors = []
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:  # 논리명이 없으면 건너뜀
                continue
            record = StdWord(
                logical_name=str(row[0]).strip(),
                physical_name=str(row[1]).strip(),
                physical_desc=_clean_str(row[2]),
                is_class_word=_parse_bool(row[3]),
                synonym=_clean_str(row[4]),
                description=_clean_str(row[5]),
            )
            batch.append(record)
            success_count += 1

            if len(batch) >= BATCH_SIZE:
                db.add_all(batch)
                await db.flush()
                batch = []
        except Exception as e:
            error_count += 1
            errors.append({"row": i, "error": str(e)})

    if batch:
        db.add_all(batch)

    await db.commit()
    wb.close()

    return {
        "total_rows": success_count + error_count,
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],  # 최대 50건의 에러만 반환
        "message": f"단어사전 임포트 완료: {success_count}건 성공, {error_count}건 실패",
    }


async def import_domain_dict(db: AsyncSession, file_bytes: bytes) -> dict:
    """도메인사전 엑셀 임포트"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    # 도메인그룹 캐시 조회
    group_result = await db.execute(select(StdDomainGroup))
    group_map = {g.group_name: g.group_id for g in group_result.scalars().all()}

    success_count = 0
    error_count = 0
    errors = []
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:
                continue
            group_name = str(row[0]).strip()
            group_id = group_map.get(group_name)
            if not group_id:
                error_count += 1
                errors.append({"row": i, "error": f"도메인그룹 '{group_name}' 없음"})
                continue

            record = StdDomain(
                group_id=group_id,
                domain_name=str(row[1]).strip(),
                domain_logical_name=str(row[2]).strip(),
                data_type=str(row[3]).strip() if row[3] else "VARCHAR",
                data_length=_parse_int(row[4]),
                data_scale=_parse_int(row[5]),
                is_personal_info=_parse_bool(row[6]),
                is_encrypted=_parse_bool(row[7]),
                scramble_type=_clean_str(row[8]),
                description=_clean_str(row[9]),
            )
            batch.append(record)
            success_count += 1

            if len(batch) >= BATCH_SIZE:
                db.add_all(batch)
                await db.flush()
                batch = []
        except Exception as e:
            error_count += 1
            errors.append({"row": i, "error": str(e)})

    if batch:
        db.add_all(batch)

    await db.commit()
    wb.close()

    return {
        "total_rows": success_count + error_count,
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],
        "message": f"도메인사전 임포트 완료: {success_count}건 성공, {error_count}건 실패",
    }


async def import_term_dict(db: AsyncSession, file_bytes: bytes) -> dict:
    """용어사전 엑셀 임포트"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    # 도메인그룹 캐시
    group_result = await db.execute(select(StdDomainGroup))
    group_map = {g.group_name: g.group_id for g in group_result.scalars().all()}

    success_count = 0
    error_count = 0
    errors = []
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:
                continue

            domain_group_name = _clean_str(row[5])
            domain_group_id = group_map.get(domain_group_name) if domain_group_name else None

            record = StdTerm(
                logical_name=str(row[0]).strip(),
                physical_name=str(row[1]).strip(),
                english_name=_clean_str(row[2]),
                domain_logical_name=_clean_str(row[3]),
                domain_abbr=_clean_str(row[4]),
                domain_group_id=domain_group_id,
                data_type=_clean_str(row[6]),
                data_length=_parse_int(row[7]),
                data_scale=_parse_int(row[8]),
                is_personal_info=_parse_bool(row[9]),
                is_encrypted=_parse_bool(row[10]),
                scramble_type=_clean_str(row[11]),
                description=_clean_str(row[12]),
            )
            batch.append(record)
            success_count += 1

            if len(batch) >= BATCH_SIZE:
                db.add_all(batch)
                await db.flush()
                batch = []
        except Exception as e:
            error_count += 1
            errors.append({"row": i, "error": str(e)})

    if batch:
        db.add_all(batch)

    await db.commit()
    wb.close()

    return {
        "total_rows": success_count + error_count,
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],
        "message": f"용어사전 임포트 완료: {success_count}건 성공, {error_count}건 실패",
    }


async def export_dict(db: AsyncSession, dict_type: str) -> StreamingResponse:
    """사전 엑셀 엑스포트"""
    wb = Workbook()
    ws = wb.active

    if dict_type == "word":
        ws.title = "단어사전"
        ws.append(["논리명", "물리명", "물리의미", "속성분류어", "동의어", "설명", "상태"])
        result = await db.execute(select(StdWord).where(StdWord.status == "active").order_by(StdWord.word_id))
        for w in result.scalars().all():
            ws.append([w.logical_name, w.physical_name, w.physical_desc,
                       "Y" if w.is_class_word else "N", w.synonym, w.description, w.status])

    elif dict_type == "domain":
        ws.title = "도메인사전"
        ws.append(["도메인그룹명", "도메인명", "도메인논리명", "데이터유형", "길이", "소수점",
                    "개인정보여부", "암호화여부", "스크램블", "설명"])
        result = await db.execute(
            select(StdDomain, StdDomainGroup.group_name)
            .join(StdDomainGroup, StdDomain.group_id == StdDomainGroup.group_id)
            .where(StdDomain.status == "active")
            .order_by(StdDomain.domain_id)
        )
        for d, gname in result.all():
            ws.append([gname, d.domain_name, d.domain_logical_name, d.data_type,
                        d.data_length, d.data_scale,
                        "Y" if d.is_personal_info else "", "Y" if d.is_encrypted else "",
                        d.scramble_type, d.description])

    elif dict_type == "term":
        ws.title = "용어사전"
        ws.append(["논리명", "물리명", "영문의미", "도메인논리명", "도메인논리약어", "도메인그룹",
                    "데이터유형", "길이", "소수점", "개인정보여부", "암호화여부", "스크램블", "설명"])
        result = await db.execute(
            select(StdTerm).where(StdTerm.status == "active").order_by(StdTerm.term_id)
        )
        for t in result.scalars().all():
            ws.append([t.logical_name, t.physical_name, t.english_name,
                        t.domain_logical_name, t.domain_abbr, None,
                        t.data_type, t.data_length, t.data_scale,
                        "Y" if t.is_personal_info else "N", "Y" if t.is_encrypted else "N",
                        t.scramble_type, t.description])

    # 엑셀을 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    filename = f"meta_{dict_type}_export.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
